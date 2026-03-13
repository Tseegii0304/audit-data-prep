"""
Аудитын эрсдэл илрүүлэх хиймэл оюуны систем v5.1
TB + Ledger + Part1 + ML + XAI → Тусдаа шинжилгээний цэс
pip install streamlit pandas numpy scikit-learn plotly openpyxl shap
streamlit run audit_app.py
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from sklearn.ensemble import IsolationForest, RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.neighbors import LocalOutlierFactor
from sklearn.svm import OneClassSVM
from sklearn.cluster import KMeans
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.metrics import precision_score, recall_score, f1_score, roc_auc_score, roc_curve
import warnings, io, re, gzip, zipfile
from datetime import datetime
from collections import Counter
warnings.filterwarnings('ignore')
from tab_descriptions import TabDescriptions
try:
    import shap
except Exception:
    shap = None

td = TabDescriptions()
st.set_page_config(page_title="Аудитын эрсдэл илрүүлэх систем v5.1", page_icon="🔍", layout="wide")
st.markdown('<h1 style="text-align:center;color:#1565c0">🔍 Аудитын эрсдэл илрүүлэх хиймэл оюуны систем</h1>', unsafe_allow_html=True)
st.markdown('<p style="text-align:center;color:#666;font-size:14px">Гүйлгээ-баланс + Ерөнхий журнал → Дансны болон гүйлгээний түвшний эрсдэл илрүүлэлт • Материаллаг байдлын тооцоо</p>', unsafe_allow_html=True)


# ── Session state defaults ──
SESSION_DEFAULTS = {
    'tb_analysis_done': False,
    'journal_ai_done': False,
    'tb_detected_rows': [],
    'journal_detected_rows': [],
    'tb_upload_cache': {},
    'journal_upload_cache': {},
}
for _k, _v in SESSION_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

def uploaded_files_to_cache(files):
    cache = {}
    for f in files or []:
        try:
            raw = f.getvalue()
        except Exception:
            raw = f.read()
            f.seek(0)
        cache[f.name] = raw
    return cache

def cache_to_file_objects(cache):
    objs = []
    for name, raw in (cache or {}).items():
        bio = io.BytesIO(raw)
        bio.name = name
        objs.append(bio)
    return objs

with st.sidebar:
    st.header("📌 Системийн цэс")
    page = st.radio("Үндсэн цэс:", ["1️⃣ Өгөгдөл оруулах, бэлтгэх", "2️⃣ Гүйлгээ балансын шинжилгээ", "3️⃣ Ерөнхий журналын шинжилгээ", "4️⃣ Материаллаг байдлын тооцоо"])

ACCT_RE_B = re.compile(r'Данс:\s*\[([^\]]+)\]\s*(.*)')
ACCT_RE_P = re.compile(r'Данс:\s*(\d{3}-\d{2}-\d{2}-\d{3})\s+(.*)')

def parse_account(text):
    m = ACCT_RE_B.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    m = ACCT_RE_P.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, None

def safe_float(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0

def process_raw_tb(file_obj):
    import openpyxl
    base_cols = ['account_code','account_name','opening_debit','opening_credit','turnover_debit','turnover_credit','closing_debit','closing_credit']
    try:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            try:
                int(float(row[0]))
            except Exception:
                continue
            code = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if not code or not re.match(r'\d{3}-', code):
                continue
            rows.append({
                'account_code': code,
                'account_name': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                'opening_debit': safe_float(row[3]) if len(row) > 3 else 0.0,
                'opening_credit': safe_float(row[4]) if len(row) > 4 else 0.0,
                'turnover_debit': safe_float(row[5]) if len(row) > 5 else 0.0,
                'turnover_credit': safe_float(row[6]) if len(row) > 6 else 0.0,
                'closing_debit': safe_float(row[7]) if len(row) > 7 else 0.0,
                'closing_credit': safe_float(row[8]) if len(row) > 8 else 0.0,
            })
        wb.close()
    except Exception:
        rows = []

    if not rows:
        empty_df = pd.DataFrame(columns=base_cols + ['opening_balance_signed','turnover_net_signed','closing_balance_signed','net_change_signed'])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as w:
            empty_df[base_cols].to_excel(w, sheet_name='01_TB_CLEAN', index=False)
            empty_df.to_excel(w, sheet_name='02_ACCOUNT_SUMMARY', index=False)
        buf.seek(0)
        return buf, empty_df

    df = pd.DataFrame(rows, columns=base_cols)
    for c in base_cols[2:]:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)

    df['opening_balance_signed'] = df['opening_debit'] - df['opening_credit']
    df['turnover_net_signed'] = df['turnover_debit'] - df['turnover_credit']
    df['closing_balance_signed'] = df['closing_debit'] - df['closing_credit']
    df['net_change_signed'] = df['closing_balance_signed'] - df['opening_balance_signed']
    tb_sum = df[['account_code','account_name','opening_debit','opening_credit','opening_balance_signed',
                  'turnover_debit','turnover_credit','turnover_net_signed',
                  'closing_debit','closing_credit','closing_balance_signed','net_change_signed']].copy()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df[base_cols].to_excel(w, sheet_name='01_TB_CLEAN', index=False)
        tb_sum.to_excel(w, sheet_name='02_ACCOUNT_SUMMARY', index=False)
    buf.seek(0)
    return buf, tb_sum

COL_PATTERNS = {
    'account_code': ['дансны код','данс код','account code','account no','account number','acc code','код'],
    'account_name': ['дансны нэр','данс нэр','account name','acc name','нэр'],
    'transaction_date': ['огноо','date','transaction date','txn date'],
    'debit_mnt': ['дебит','debit','dt','дт','debit amount'],
    'credit_mnt': ['кредит','credit','ct','кт','credit amount'],
    'balance_mnt': ['үлдэгдэл','balance','bal','ending balance'],
    'counterparty_name': ['харилцагч','counterparty','partner','vendor','customer'],
    'transaction_description': ['тайлбар','гүйлгээний утга','утга','description','memo','narration'],
    'journal_no': ['журнал','journal','journal no'],
    'document_no': ['баримт','document','doc no'],
}
def _match_col(h, field):
    h2 = str(h).lower().strip()
    return any(p in h2 for p in COL_PATTERNS.get(field, []))
def _auto_map(headers):
    m, used = {}, set()
    for f in ['account_code','debit_mnt','credit_mnt','transaction_date','account_name','counterparty_name','transaction_description','balance_mnt','journal_no','document_no']:
        for i, h in enumerate(headers):
            if i in used: continue
            if _match_col(h, f): m[f]=i; used.add(i); break
    return m

def _find_header_row(all_rows, max_scan=30):
    """Гарчигын мөрийг автоматаар хайна."""
    best_i, best_s = 0, 0
    for i, row in enumerate(all_rows[:max_scan]):
        vals = [str(c).strip().lower() for c in row if c is not None]
        score = 0
        for v in vals:
            if 'огноо' in v or 'date' in v: score += 1
            if 'дебет' in v or 'debit' in v: score += 1
            if 'кредит' in v or 'credit' in v: score += 1
            if 'мөнгөн дүн' in v or 'amount' in v: score += 1
            if 'гүйлгээний утга' in v or 'description' in v: score += 1
            if 'код' in v or 'code' in v or 'account' in v: score += 1
            if 'баримт' in v or 'document' in v: score += 1
            if 'журнал' in v or 'journal' in v: score += 1
        if score > best_s:
            best_s, best_i = score, i
    return best_i, best_s

def process_edt(file_obj, report_year):
    """Янз бүрийн ЕЖ / ерөнхий журнал форматыг бүх sheet-ээр шалгаж уншина.
    3 parser: standard (Данс:[...]), dual-entry (дебет/кредит данс), rowwise (мөр бүрд данс+дүн)
    """
    import openpyxl
    EDT_COLUMNS = ['report_year','account_code','account_name','transaction_no','transaction_date',
                   'journal_no','document_no','counterparty_name','counterparty_id',
                   'transaction_description','debit_mnt','credit_mnt','balance_mnt','month']

    def _to_date(v):
        if v is None: return ''
        if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
        s = str(v).strip()
        for fmt in ('%Y-%m-%d','%Y/%m/%d','%Y.%m.%d','%d.%m.%Y','%y.%m.%d','%y-%m-%d'):
            try: return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except: pass
        m = re.match(r'^(\d{2})[./-](\d{2})[./-](\d{2})$', s)
        if m:
            yy, mm, dd = m.groups()
            return f'20{yy}-{mm}-{dd}'
        return s[:10]

    def _pick(headers, candidates):
        headers_l = [str(h).strip().lower() if h is not None else '' for h in headers]
        for cand in candidates:
            for i, h in enumerate(headers_l):
                if cand in h: return i
        return None

    # ═══ Parser 1: Стандарт ЕЖ (Данс: [...]) ═══
    def _parse_standard_sheet(ws):
        rows_out, cur_code, cur_name = [], None, None
        for row in ws.iter_rows(values_only=True):
            c0 = row[0] if len(row) > 0 else None
            if c0 is None: continue
            s = str(c0).strip()
            if s.startswith('Данс:'):
                code, name = parse_account(s)
                if code: cur_code, cur_name = code, name
                continue
            if any(s.startswith(x) for x in ['Компани:','ЕРӨНХИЙ','Тайлант','Үүсгэсэн','Журнал:','№','Эцсийн','Дт -','Нийт','Эхний','Нээгээд']) or s in ('Валютаар','Төгрөгөөр',''):
                continue
            try: tx_no = int(float(c0))
            except: continue
            if cur_code is None: continue
            tx_date = _to_date(row[1] if len(row) > 1 else '')
            rows_out.append({'report_year':str(report_year),'account_code':cur_code,'account_name':cur_name,
                'transaction_no':str(tx_no),'transaction_date':tx_date,
                'journal_no':str(row[5]).strip() if len(row)>5 and row[5] else '',
                'document_no':str(row[6]).strip() if len(row)>6 and row[6] else '',
                'counterparty_name':str(row[3]).strip() if len(row)>3 and row[3] else '',
                'counterparty_id':str(row[4]).strip() if len(row)>4 and row[4] else '',
                'transaction_description':str(row[7]).strip() if len(row)>7 and row[7] else '',
                'debit_mnt':safe_float(row[9]) if len(row)>9 else 0.0,
                'credit_mnt':safe_float(row[11]) if len(row)>11 else 0.0,
                'balance_mnt':safe_float(row[13]) if len(row)>13 else 0.0,
                'month':tx_date[:7] if len(tx_date)>=7 else ''})
        return pd.DataFrame(rows_out, columns=EDT_COLUMNS), len(rows_out)

    # ═══ Parser 2: Dual-entry журнал (дебет данс + кредит данс + дүн) ═══
    def _parse_dual_entry_sheet(ws):
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(list(row))
            if i >= 5000: break
        if not all_rows: return pd.DataFrame(columns=EDT_COLUMNS), 0
        hdr_i, hdr_score = _find_header_row(all_rows)
        if hdr_score < 3: return pd.DataFrame(columns=EDT_COLUMNS), 0
        headers = [str(c).strip() if c is not None else f'col_{j}' for j, c in enumerate(all_rows[hdr_i])]
        debit_idx = _pick(headers, ['дебет', 'debit'])
        credit_idx = _pick(headers, ['кредит', 'credit'])
        amount_idx = _pick(headers, ['мөнгөн дүн', 'amount', 'дүн'])
        date_idx = _pick(headers, ['огноо', 'date'])
        doc_idx = _pick(headers, ['баримт №', 'баримт', 'document', 'doc'])
        cp_idx = _pick(headers, ['байгууллагын нэр', 'харилцагч', 'counterparty', 'customer', 'vendor'])
        desc_idx = _pick(headers, ['гүйлгээний утга', 'тайлбар', 'description', 'memo'])
        j_idx = _pick(headers, ['журналын төрөл', 'журнал', 'journal'])
        if debit_idx is None or credit_idx is None or amount_idx is None:
            return pd.DataFrame(columns=EDT_COLUMNS), 0
        rows_out = []
        tx_counter = 0
        for row in all_rows[hdr_i+1:]:
            if not row or all(c is None or str(c).strip()=='' for c in row): continue
            debit_acct = str(row[debit_idx]).strip() if debit_idx < len(row) and row[debit_idx] is not None else ''
            credit_acct = str(row[credit_idx]).strip() if credit_idx < len(row) and row[credit_idx] is not None else ''
            amount = safe_float(row[amount_idx]) if amount_idx < len(row) else 0.0
            if not re.search(r'\d', debit_acct or '') or not re.search(r'\d', credit_acct or '') or amount == 0: continue
            tx_date = _to_date(row[date_idx]) if date_idx is not None and date_idx < len(row) else ''
            doc_no = str(row[doc_idx]).strip() if doc_idx is not None and doc_idx < len(row) and row[doc_idx] is not None else ''
            cp_name = str(row[cp_idx]).strip() if cp_idx is not None and cp_idx < len(row) and row[cp_idx] is not None else ''
            desc = str(row[desc_idx]).strip() if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None else ''
            journal_no = str(row[j_idx]).strip() if j_idx is not None and j_idx < len(row) and row[j_idx] is not None else ''
            tx_counter += 1
            common = {'report_year':str(report_year),'transaction_no':str(tx_counter),'transaction_date':tx_date,
                      'journal_no':journal_no,'document_no':doc_no,'counterparty_name':cp_name,'counterparty_id':'',
                      'transaction_description':desc,'balance_mnt':0.0,'month':tx_date[:7] if len(tx_date)>=7 else ''}
            rows_out.append({**common, 'account_code':debit_acct, 'account_name':'', 'debit_mnt':amount, 'credit_mnt':0.0})
            rows_out.append({**common, 'account_code':credit_acct, 'account_name':'', 'debit_mnt':0.0, 'credit_mnt':amount})
        return pd.DataFrame(rows_out, columns=EDT_COLUMNS), len(rows_out)

    # ═══ Parser 3: Rowwise журнал (мөр бүрд данс код + дебит + кредит) ═══
    def _parse_rowwise_sheet(ws):
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(list(row))
            if i >= 5000: break
        if not all_rows: return pd.DataFrame(columns=EDT_COLUMNS), 0
        hdr_i, hdr_score = _find_header_row(all_rows)
        if hdr_score < 3: return pd.DataFrame(columns=EDT_COLUMNS), 0
        headers = [str(c).strip() if c is not None else f'col_{j}' for j, c in enumerate(all_rows[hdr_i])]
        date_idx = _pick(headers, ['огноо', 'date'])
        doc_idx = _pick(headers, ['баримтын дугаар', 'баримт №', 'баримт', 'document'])
        code_idx = _pick(headers, ['код', 'account', 'данс'])
        name_idx = None
        if code_idx is not None and code_idx + 1 < len(headers):
            name_idx = code_idx + 1
        debit_amt_idx = _pick(headers, ['sum of дебет', 'debit amount', 'дебет', 'дт', 'debit'])
        credit_amt_idx = _pick(headers, ['sum of кредит', 'credit amount', 'кредит', 'кт', 'credit'])
        desc_idx = _pick(headers, ['гүйлгээний утга', 'тайлбар', 'description'])
        journal_idx = _pick(headers, ['журнал', 'journal'])
        cp_idx = _pick(headers, ['харилцагч', 'байгууллагын нэр', 'vendor', 'customer'])
        if code_idx is None or (debit_amt_idx is None and credit_amt_idx is None):
            return pd.DataFrame(columns=EDT_COLUMNS), 0
        rows_out = []
        for idx, row in enumerate(all_rows[hdr_i+1:], start=1):
            if not row or all(c is None or str(c).strip()=='' for c in row): continue
            acct = str(row[code_idx]).strip() if code_idx < len(row) and row[code_idx] is not None else ''
            if not re.search(r'\d', acct or ''): continue
            db = safe_float(row[debit_amt_idx]) if debit_amt_idx is not None and debit_amt_idx < len(row) else 0.0
            cr = safe_float(row[credit_amt_idx]) if credit_amt_idx is not None and credit_amt_idx < len(row) else 0.0
            if db == 0 and cr == 0: continue
            tx_date = _to_date(row[date_idx]) if date_idx is not None and date_idx < len(row) else ''
            doc_no = str(row[doc_idx]).strip() if doc_idx is not None and doc_idx < len(row) and row[doc_idx] is not None else ''
            acct_name = str(row[name_idx]).strip() if name_idx is not None and name_idx < len(row) and row[name_idx] is not None else ''
            cp_name = str(row[cp_idx]).strip() if cp_idx is not None and cp_idx < len(row) and row[cp_idx] is not None else ''
            desc = str(row[desc_idx]).strip() if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None else ''
            journal_no = str(row[journal_idx]).strip() if journal_idx is not None and journal_idx < len(row) and row[journal_idx] is not None else ''
            rows_out.append({'report_year':str(report_year),'account_code':acct,'account_name':acct_name,'transaction_no':str(idx),
                'transaction_date':tx_date,'journal_no':journal_no,'document_no':doc_no,'counterparty_name':cp_name,
                'counterparty_id':'','transaction_description':desc,'debit_mnt':db,'credit_mnt':cr,'balance_mnt':0.0,
                'month':tx_date[:7] if len(tx_date)>=7 else ''})
        return pd.DataFrame(rows_out, columns=EDT_COLUMNS), len(rows_out)

    # ═══ Parser 4: Монголын ерөнхий журнал (Д/д | Огноо | Дугаар | Утга | Данс | Дебет | Кредит) ═══
    def _parse_mongolian_journal(ws):
        """Монголын стандарт ерөнхий журнал формат:
        Row ~8: Д/д | Баримтын | | Гүйлгээний утга | Харьцсан данс | Дүн |
        Row ~9: | Огноо | Дугаар | | | Дебет | Кредит
        Data: seq | date | doc | desc | account | debit | credit
        """
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(list(row))
            if i >= 10000: break
        if len(all_rows) < 10:
            return pd.DataFrame(columns=EDT_COLUMNS), 0

        # Гарчиг олох: "Д/д" эсвэл "Дебет" + "Кредит" агуулсан мөрийг хайна
        data_start = None
        for i in range(min(20, len(all_rows))):
            row_text = ' '.join(str(c).strip().lower() for c in all_rows[i] if c is not None)
            if 'д/д' in row_text or ('дебет' in row_text and 'кредит' in row_text):
                # Дараагийн мөр нь "Огноо" "Дугаар" агуулж магадгүй (2 мөрт гарчиг)
                if i + 1 < len(all_rows):
                    next_text = ' '.join(str(c).strip().lower() for c in all_rows[i+1] if c is not None)
                    if 'огноо' in next_text or 'дугаар' in next_text:
                        data_start = i + 2
                    else:
                        data_start = i + 1
                else:
                    data_start = i + 1
                break

        if data_start is None:
            return pd.DataFrame(columns=EDT_COLUMNS), 0

        # Дебет, кредит баганын индексийг олох
        # Гарчигын мөрүүдээс "Дебет", "Кредит" хайна
        debit_col, credit_col = None, None
        for check_row in range(max(0, data_start - 3), data_start):
            for j, cell in enumerate(all_rows[check_row]):
                if cell is None: continue
                cl = str(cell).strip().lower()
                if cl in ('дебет', 'дебит', 'debit', 'дүн') and debit_col is None:
                    debit_col = j
                elif cl in ('кредит', 'кредит', 'credit') and credit_col is None:
                    credit_col = j

        # Хэрэв олдохгүй бол 7 баганатай бол F=5, G=6 гэж таамаглана
        if debit_col is None or credit_col is None:
            max_cols = max(len(r) for r in all_rows[data_start:data_start+5]) if all_rows[data_start:data_start+5] else 0
            if max_cols >= 7:
                debit_col = debit_col or 5
                credit_col = credit_col or 6
            else:
                return pd.DataFrame(columns=EDT_COLUMNS), 0

        rows_out = []
        for row in all_rows[data_start:]:
            if not row or all(c is None or str(c).strip() == '' for c in row): continue
            # A=seq, B=date, C=doc, D=desc, E=account, F=debit, G=credit
            c0 = row[0] if len(row) > 0 else None
            if c0 is None: continue
            # Д/д нь тоо байх ёстой
            try:
                int(float(c0))
            except:
                # "Нийт", "Дүн" гэх мэт мөрийг алгасна
                s0 = str(c0).strip()
                if any(s0.startswith(x) for x in ['Нийт','Дүн','Журнал','Ерөнхий','Бүгд']):
                    continue
                continue

            tx_date = _to_date(row[1] if len(row) > 1 else '')
            doc_no = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            desc = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            acct = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            db = safe_float(row[debit_col]) if debit_col < len(row) else 0.0
            cr = safe_float(row[credit_col]) if credit_col < len(row) else 0.0

            if db == 0 and cr == 0: continue
            if not acct or acct in ('None','nan',''): acct = '000'

            rows_out.append({
                'report_year': str(report_year), 'account_code': acct, 'account_name': '',
                'transaction_no': str(len(rows_out) + 1), 'transaction_date': tx_date,
                'journal_no': '', 'document_no': doc_no,
                'counterparty_name': '', 'counterparty_id': '',
                'transaction_description': desc,
                'debit_mnt': db, 'credit_mnt': cr, 'balance_mnt': 0.0,
                'month': tx_date[:7] if len(tx_date) >= 7 else ''
            })
        return pd.DataFrame(rows_out, columns=EDT_COLUMNS), len(rows_out)

    # ═══ Бүх sheet, бүх parser-ийг оролдоно ═══
    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        best_df = pd.DataFrame(columns=EDT_COLUMNS)
        best_cnt = 0
        for sname in wb.sheetnames:
            ws = wb[sname]
            for parser in (_parse_mongolian_journal, _parse_standard_sheet, _parse_dual_entry_sheet, _parse_rowwise_sheet):
                try:
                    df_try, cnt_try = parser(ws)
                except Exception:
                    continue
                if cnt_try > best_cnt:
                    best_df, best_cnt = df_try, cnt_try
        wb.close()
        return best_df, best_cnt
    except Exception:
        return pd.DataFrame(columns=EDT_COLUMNS), 0

def generate_part1(df_led, year):
    df = df_led.copy()
    yr = str(year)
    df['debit_mnt'] = pd.to_numeric(df['debit_mnt'], errors='coerce').fillna(0)
    df['credit_mnt'] = pd.to_numeric(df['credit_mnt'], errors='coerce').fillna(0)
    df['balance_mnt'] = pd.to_numeric(df['balance_mnt'], errors='coerce').fillna(0)
    monthly = df.groupby(['month', 'account_code']).agg(
        total_debit_mnt=('debit_mnt', 'sum'),
        total_credit_mnt=('credit_mnt', 'sum'),
        ending_balance_mnt=('balance_mnt', 'last'),
        transaction_count=('debit_mnt', 'count')
    ).reset_index()
    monthly.insert(0, 'report_year', yr)
    anames = df.groupby('account_code')['account_name'].first()
    acct = df.groupby('account_code').agg(
        total_debit_mnt=('debit_mnt', 'sum'),
        total_credit_mnt=('credit_mnt', 'sum'),
        closing_balance_mnt=('balance_mnt', 'last')
    ).reset_index()
    acct['account_name'] = acct['account_code'].map(anames)
    acct.insert(0, 'report_year', yr)
    rm = df.groupby(['month', 'account_code', 'counterparty_name']).agg(
        transaction_count=('debit_mnt', 'count'),
        total_debit=('debit_mnt', 'sum'),
        total_credit=('credit_mnt', 'sum'),
    ).reset_index()
    rm['total_amount_mnt'] = rm['total_debit'].abs() + rm['total_credit'].abs()
    rm.insert(0, 'report_year', yr)
    p75a = rm['total_amount_mnt'].quantile(0.75)
    p75c = rm['transaction_count'].quantile(0.75)
    rm['risk_flag_large_txn'] = (rm['total_amount_mnt'] > p75a).astype(int)
    rm['risk_flag_high_frequency'] = (rm['transaction_count'] > p75c).astype(int)
    rm['risk_score'] = rm['risk_flag_large_txn'] + rm['risk_flag_high_frequency']
    rm['risk_level'] = pd.cut(
        rm['risk_score'],
        bins=[-0.1, 0.5, 1.5, 99],
        labels=['Бага', 'Дунд', 'Өндөр']
    ).astype(str)
    rm['account_category'] = rm['account_code'].str[:1].map(
        {'1': 'Хөрөнгө', '2': 'Өр', '3': 'Эздийн өмч', '4': 'Зардал', '5': 'Орлого', '6': 'Орлого', '7': 'Зардал'}
    ).fillna('')
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        monthly.to_excel(w, sheet_name='02_MONTHLY_SUMMARY', index=False)
        acct.to_excel(w, sheet_name='03_ACCOUNT_SUMMARY', index=False)
        rm.to_excel(w, sheet_name='04_RISK_MATRIX', index=False)
    buf.seek(0)
    n_risk = len(rm[rm['risk_score'] > 0])
    return buf, monthly, acct, rm, n_risk

def read_ledger(f):
    raw = f.read()
    f.seek(0)
    if raw[:2] == b'\x1f\x8b':
        return pd.read_csv(io.StringIO(gzip.decompress(raw).decode('utf-8')), dtype={'account_code': str})
    return pd.read_csv(io.BytesIO(raw), dtype={'account_code': str})

def get_year(name):
    for y in range(2020, 2030):
        if str(y) in name:
            return y
    return 2025

def load_tb(files):
    frames = []
    stats = {}
    must_cols = ['account_code','account_name','opening_debit','opening_credit','opening_balance_signed',
                 'turnover_debit','turnover_credit','turnover_net_signed',
                 'closing_debit','closing_credit','closing_balance_signed','net_change_signed']
    for f in files:
        year = get_year(f.name)
        try:
            df = pd.read_excel(f, sheet_name='02_ACCOUNT_SUMMARY')
        except Exception:
            try:
                f.seek(0)
                df = pd.read_excel(f)
            except Exception:
                df = pd.DataFrame()
        if df.empty:
            continue
        for c in must_cols:
            if c not in df.columns:
                if c in ['account_code','account_name']:
                    df[c] = ''
                else:
                    df[c] = 0.0
        df['year'] = year
        for c in ['turnover_debit', 'turnover_credit', 'closing_debit', 'closing_credit', 'opening_debit', 'opening_credit', 'net_change_signed']:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        stats[year] = {'accounts': len(df), 'turnover_d': df['turnover_debit'].sum(), 'turnover_c': df['turnover_credit'].sum()}
        frames.append(df[must_cols + ['year']])
    if not frames:
        return pd.DataFrame(columns=must_cols + ['year']), {}
    return pd.concat(frames, ignore_index=True), stats


def load_ledger_stats(files, sample_per_year=20000, chunksize=100000):
    """Ledger файлуудыг chunk-ээр уншиж stats + sample DataFrame буцаана.
    Өндөр хэмжээтэй ledger дээр Streamlit Cloud OOM болохоос сэргийлнэ.
    """
    stats = {}
    sampled_frames = []
    needed_cols = [
        'report_year','account_code','account_name','transaction_no','transaction_date',
        'journal_no','document_no','counterparty_name','counterparty_id',
        'transaction_description','debit_mnt','credit_mnt','balance_mnt','month'
    ]

    def _iter_chunks(fobj):
        fobj.seek(0)
        raw = fobj.read()
        fobj.seek(0)
        if raw[:2] == b'\x1f\x8b':
            bio = io.BytesIO(gzip.decompress(raw))
            return pd.read_csv(bio, dtype={'account_code': str}, chunksize=chunksize)
        return pd.read_csv(io.BytesIO(raw), dtype={'account_code': str}, chunksize=chunksize)

    for f in files:
        year = get_year(f.name)
        total_rows = 0
        acct_set = set()
        month_set = set()
        monthly_parts = []
        year_samples = []

        try:
            for chunk in _iter_chunks(f):
                total_rows += len(chunk)

                for c in needed_cols:
                    if c not in chunk.columns:
                        chunk[c] = '' if c in (
                            'report_year','account_code','account_name','transaction_no','transaction_date',
                            'journal_no','document_no','counterparty_name','counterparty_id',
                            'transaction_description','month'
                        ) else 0

                chunk['account_code'] = chunk['account_code'].astype(str)
                chunk['debit_mnt'] = pd.to_numeric(chunk['debit_mnt'], errors='coerce').fillna(0)
                chunk['credit_mnt'] = pd.to_numeric(chunk['credit_mnt'], errors='coerce').fillna(0)
                chunk['report_year'] = str(year)

                acct_set.update(chunk['account_code'].dropna().astype(str).unique().tolist())
                month_set.update(chunk['month'].dropna().astype(str).unique().tolist())

                mo = chunk.groupby('month').agg(
                    rows=('debit_mnt', 'count'),
                    debit=('debit_mnt', 'sum'),
                    credit=('credit_mnt', 'sum')
                ).reset_index()
                monthly_parts.append(mo)

                # sample cap
                current_n = sum(len(x) for x in year_samples)
                remain = max(sample_per_year - current_n, 0)
                if remain > 0:
                    take_n = min(len(chunk), max(1000, remain))
                    year_samples.append(chunk.sample(n=min(take_n, len(chunk)), random_state=42)[needed_cols].copy())

        except Exception:
            # fallback жижиг файл дээр full read
            f.seek(0)
            df = read_ledger(f)
            total_rows = len(df)
            for c in needed_cols:
                if c not in df.columns:
                    df[c] = '' if c in (
                        'report_year','account_code','account_name','transaction_no','transaction_date',
                        'journal_no','document_no','counterparty_name','counterparty_id',
                        'transaction_description','month'
                    ) else 0
            df['account_code'] = df['account_code'].astype(str)
            df['debit_mnt'] = pd.to_numeric(df['debit_mnt'], errors='coerce').fillna(0)
            df['credit_mnt'] = pd.to_numeric(df['credit_mnt'], errors='coerce').fillna(0)
            df['report_year'] = str(year)
            acct_set.update(df['account_code'].dropna().astype(str).unique().tolist())
            month_set.update(df['month'].dropna().astype(str).unique().tolist())
            monthly_parts.append(df.groupby('month').agg(rows=('debit_mnt', 'count'), debit=('debit_mnt', 'sum'), credit=('credit_mnt', 'sum')).reset_index())
            year_samples.append(df.sample(n=min(sample_per_year, len(df)), random_state=42)[needed_cols].copy())

        mo = pd.concat(monthly_parts, ignore_index=True).groupby('month').agg(
            rows=('rows', 'sum'),
            debit=('debit', 'sum'),
            credit=('credit', 'sum')
        ).sort_index() if monthly_parts else pd.DataFrame(columns=['rows','debit','credit'])

        stats[year] = {
            'rows': int(total_rows),
            'accounts': int(len(acct_set)),
            'months': int(len(month_set)),
            'monthly': mo
        }

        if year_samples:
            year_sample = pd.concat(year_samples, ignore_index=True).head(sample_per_year)
            year_sample['report_year'] = str(year)
            sampled_frames.append(year_sample)

    full_df = pd.concat(sampled_frames, ignore_index=True) if sampled_frames else pd.DataFrame(columns=needed_cols)
    return stats, full_df

def load_part1(files):
    all_rm = []
    all_mo = []
    for f in files:
        year = get_year(f.name)
        try:
            rm = pd.read_excel(f, sheet_name='04_RISK_MATRIX')
            rm['year'] = year
            all_rm.append(rm)
        except Exception:
            pass
        try:
            mo = pd.read_excel(f, sheet_name='02_MONTHLY_SUMMARY')
            mo['year'] = year
            all_mo.append(mo)
        except Exception:
            pass
    rm_all = pd.concat(all_rm, ignore_index=True) if all_rm else pd.DataFrame()
    mo_all = pd.concat(all_mo, ignore_index=True) if all_mo else pd.DataFrame()
    return rm_all, mo_all

def clean_for_risk(df):
    """NaN / хоосон утгыг цэвэрлэж, эрсдэлийн шинжилгээнд бэлтгэнэ."""
    d = df.copy()
    text_cols = ['account_code','account_name','counterparty_name','transaction_description','journal_no','document_no']
    for c in text_cols:
        if c in d.columns:
            d[c] = d[c].astype(str).replace(['nan','None','NaN','<NA>'],'').fillna('')
            d[c] = d[c].replace(r'^\s*$', '', regex=True)
    for c in ['debit_mnt','credit_mnt','balance_mnt']:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors='coerce').fillna(0.0)
    if 'account_code' in d.columns:
        d = d[d['account_code'].astype(str).str.strip() != '']
    return d.reset_index(drop=True)


def engineer_txn_features(d):
    """Гүйлгээ бүрээс шинж чанар үүсгэнэ. Дутуу багана байвал 0 утга ашиглана."""
    d = d.copy()
    # Баганууд байгаа эсэхийг шалгаж, дутууг нэмэх
    for c in ['debit_mnt','credit_mnt','account_code','account_name','counterparty_name','transaction_description','transaction_date']:
        if c not in d.columns:
            d[c] = '' if c in ('account_code','account_name','counterparty_name','transaction_description','transaction_date') else 0
    d['debit_mnt'] = pd.to_numeric(d['debit_mnt'], errors='coerce').fillna(0)
    d['credit_mnt'] = pd.to_numeric(d['credit_mnt'], errors='coerce').fillna(0)
    d['account_code'] = d['account_code'].astype(str).fillna('000')
    d['account_name'] = d['account_name'].astype(str).fillna('')
    d['counterparty_name'] = d['counterparty_name'].astype(str).fillna('')
    d['transaction_description'] = d['transaction_description'].astype(str).fillna('')
    d['transaction_date'] = d['transaction_date'].astype(str).fillna('')

    d['amount'] = d['debit_mnt'].abs() + d['credit_mnt'].abs()
    d['log_amount'] = np.log1p(d['amount'])
    d['is_debit'] = (d['debit_mnt'] > 0).astype(int)

    # Дансны ангилал
    try:
        le2 = LabelEncoder()
        d['acct_cat_num'] = le2.fit_transform(d['account_code'].str[:3])
    except:
        d['acct_cat_num'] = 0

    # Бенфорд
    digits = d['amount'].apply(lambda x: int(str(int(abs(x)))[0]) if abs(x) >= 1 else 0)
    d['benford_digit'] = digits
    benford_exp = {1:0.301,2:0.176,3:0.125,4:0.097,5:0.079,6:0.067,7:0.058,8:0.051,9:0.046}
    af = d[d['benford_digit']>0]['benford_digit'].value_counts(normalize=True)
    d['benford_dev'] = d['benford_digit'].map(lambda x: abs(af.get(x,0)-benford_exp.get(x,0)) if x>0 else 0)

    # Тэгш тоо
    d['is_round'] = (((d['amount']>=1e6)&(d['amount']%1e6==0)).astype(int) + ((d['amount']>=1e3)&(d['amount']%1e3==0)).astype(int))

    # Данс доторх z-score
    try:
        as2 = d.groupby('account_code')['amount'].agg(['mean','std']).fillna(0)
        as2.columns = ['acct_mean','acct_std']
        d = d.merge(as2, on='account_code', how='left')
        d['amt_zscore'] = np.where(d['acct_std']>0, (d['amount']-d['acct_mean'])/d['acct_std'], 0)
        d['amt_zscore'] = d['amt_zscore'].clip(-10,10).fillna(0)
    except:
        d['acct_mean'] = 0; d['acct_std'] = 0; d['amt_zscore'] = 0

    # Ховор харилцагч
    try:
        cp_f = d['counterparty_name'].value_counts()
        d['cp_rare'] = (d['counterparty_name'].map(cp_f).fillna(0) <= 3).astype(int)
    except:
        d['cp_rare'] = 0

    # Ховор данс-харилцагч хос
    try:
        d['pair'] = d['account_code'] + '|' + d['counterparty_name']
        pf = d['pair'].value_counts()
        d['pair_rare'] = (d['pair'].map(pf).fillna(0) <= 2).astype(int)
    except:
        d['pair_rare'] = 0

    # Тайлбар
    d['desc_empty'] = (d['transaction_description'].str.len() == 0).astype(int)

    # Давхардал
    try:
        d['dup_key'] = d['account_code'] + '|' + d['amount'].astype(str) + '|' + d['transaction_date']
        dk = d['dup_key'].value_counts()
        d['is_dup'] = (d['dup_key'].map(dk).fillna(1) > 1).astype(int)
    except:
        d['is_dup'] = 0

    # Цаг
    d['day'] = pd.to_numeric(d['transaction_date'].str[8:10], errors='coerce').fillna(15)
    d['month_num'] = pd.to_numeric(d['transaction_date'].str[5:7], errors='coerce').fillna(6)
    d['is_month_end'] = (d['day'] >= 28).astype(int)
    d['is_year_end'] = (d['month_num'] == 12).astype(int)

    # ═══ ТАЙЛБАР ↔ ДАНСНЫ НЭР ТУЛГАЛТ ═══
    d['desc_mismatch'] = 0
    d['name_no_overlap'] = 0
    d['dir_mismatch'] = 0
    try:
        stop_w = {'дансны','данс','нийт','бусад','зардал','орлого','төлбөр','хөрөнгө','тооцоо','бүртгэл','дүн','төгрөг','сая','мянга','журнал','гүйлгээ','баримт'}
        # Данс бүрийн ердийн тайлбарын үгс
        acct_words = {}
        for code in d['account_code'].unique():
            all_desc = ' '.join(d.loc[d['account_code']==code, 'transaction_description'].str.lower())
            wc = Counter(re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', all_desc))
            acct_words[code] = set(w for w,c in wc.items() if c >= 3 and len(w) >= 3)

        # Vectorized desc_mismatch
        def _check_mismatch(code, tx_desc):
            tx = str(tx_desc).lower() if tx_desc else ''
            if not tx or code not in acct_words or not acct_words[code]: return 0
            tx_words = set(re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', tx))
            return 0 if len(tx_words & acct_words[code]) > 0 else 1
        d['desc_mismatch'] = [_check_mismatch(c, t) for c, t in zip(d['account_code'], d['transaction_description'])]

        # Vectorized name_no_overlap
        def _extract_kw(text):
            if not text: return set()
            return set(w for w in re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', str(text).lower()) if w not in stop_w and len(w) >= 3)
        def _check_overlap(aname, tdesc):
            nk = _extract_kw(aname)
            dk2 = _extract_kw(tdesc)
            if not nk or not dk2: return 0
            return 0 if len(nk & dk2) > 0 else 1
        d['name_no_overlap'] = [_check_overlap(a, t) for a, t in zip(d['account_name'], d['transaction_description'])]
    except:
        pass

    # Дансны чиглэл зөрчил
    try:
        af2 = d['account_code'].str[0]
        d.loc[(af2=='1')&(d['credit_mnt']>0)&(d['debit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2=='2')&(d['debit_mnt']>0)&(d['credit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2=='5')&(d['debit_mnt']>0)&(d['credit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2.isin(['6','7','8']))&(d['credit_mnt']>0)&(d['debit_mnt']==0), 'dir_mismatch'] = 1
    except:
        pass

    return d

def run_txn_anomaly(df, cont=0.05):
    """Гүйлгээний аномали илрүүлэлт."""
    feats = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore','cp_rare','pair_rare',
             'desc_empty','is_month_end','is_year_end','is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']
    # Бүх feature багана байгаа эсэхийг шалгах
    for f in feats:
        if f not in df.columns:
            df[f] = 0
    X = df[feats].fillna(0).replace([np.inf,-np.inf], 0).astype(float)
    iso = IsolationForest(contamination=cont, random_state=42, n_estimators=200, n_jobs=1)
    df['txn_anomaly'] = (iso.fit_predict(X)==-1).astype(int)
    df['txn_score'] = -iso.score_samples(X)
    try:
        z = np.abs(StandardScaler().fit_transform(X))
        df['txn_zscore_flag'] = (z.max(axis=1)>2.5).astype(int)
    except:
        df['txn_zscore_flag'] = 0
    # Эрсдэлийн жинлэсэн оноо (ISA стандарттай нийцүүлсэн)
    df['txn_risk'] = (
        df['txn_anomaly'] * 3 +         # IF аномали (ISA 240)
        df['txn_zscore_flag'] * 2 +       # Z-score хэт хазайлт
        df['is_dup'] * 3 +               # Давхардсан гүйлгээ (ISA 240) — жин нэмсэн
        df['cp_rare'] * 1 +              # Ховор харилцагч (ISA 550)
        df['pair_rare'] * 1 +            # Ховор данс×харилцагч хос (ISA 550)
        (df['amt_zscore'].abs() > 3).astype(int) * 2 +  # Дундажаас хэт зөрсөн (ISA 520)
        df['desc_empty'] * 2 +           # Тайлбаргүй гүйлгээ (ISA 500) — жин нэмсэн
        df['desc_mismatch'] * 2 +        # Тайлбар↔данс зөрчил (ISA 500)
        df['name_no_overlap'] * 1 +      # Нэр давхцахгүй (ISA 500)
        df['dir_mismatch'] * 3 +         # Чиглэлийн зөрчил (ISA 240) — жин нэмсэн
        df.get('is_round', pd.Series(0, index=df.index)).astype(int) * 1  # Тэгш тоо
    )
    df['txn_risk_level'] = pd.cut(df['txn_risk'], bins=[-1,3,7,12,100],
        labels=['🟢 Бага','🟡 Дунд','🟠 Өндөр','🔴 Маш өндөр'])
    return df, feats


def run_txn_ml_ensemble(df, contamination=0.05, n_clusters=8):
    """Ерөнхий журналын ML ensemble: боломжит бүх хувилбарын аномали оноо."""
    d = clean_for_risk(df)
    if d is None or d.empty:
        return pd.DataFrame(), [], pd.DataFrame(), pd.DataFrame()
    d = engineer_txn_features(d)
    if len(d) < 5:
        d = d.copy()
        d['ml_iso_flag'] = 0
        d['ml_lof_flag'] = 0
        d['ml_svm_flag'] = 0
        d['ml_kmeans_flag'] = 0
        d['ml_zscore_flag'] = 0
        d['ml_vote_count'] = 0
        d['ml_anomaly_flag'] = 0
        d['ml_risk_level'] = '🟢 Бага'
        d['xai_top_feature'] = ''
        model_summary = pd.DataFrame([{'Алгоритм':'Too few rows','Илрүүлсэн аномали':0,'Хувь':0.0}])
        xai_importance = pd.DataFrame({'feature':[], 'importance':[]})
        return d, [], model_summary, xai_importance
    feat_cols = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore','cp_rare','pair_rare',
                 'desc_empty','is_month_end','is_year_end','is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']
    for c in feat_cols:
        if c not in d.columns:
            d[c] = 0
    X = d[feat_cols].fillna(0).replace([np.inf,-np.inf],0).astype(float)
    scaler = StandardScaler()
    Xs = scaler.fit_transform(X) if len(X) > 1 else X.values

    # Isolation Forest
    iso = IsolationForest(contamination=min(max(contamination,0.01),0.40), random_state=42, n_estimators=250)
    iso_pred = (iso.fit_predict(X) == -1).astype(int)
    iso_score = -iso.score_samples(X)

    # Local Outlier Factor
    n_neighbors = max(5, min(20, len(d)-1)) if len(d) > 5 else 2
    lof = LocalOutlierFactor(n_neighbors=n_neighbors, contamination=min(max(contamination,0.01),0.40))
    lof_pred_raw = lof.fit_predict(Xs)
    lof_pred = (lof_pred_raw == -1).astype(int)
    lof_score = -lof.negative_outlier_factor_

    # One-Class SVM
    svm = OneClassSVM(nu=min(max(contamination,0.01),0.40), kernel='rbf', gamma='scale')
    svm.fit(Xs)
    svm_pred = (svm.predict(Xs) == -1).astype(int)
    svm_score = -svm.score_samples(Xs)

    # KMeans distance anomaly
    k = max(2, min(n_clusters, len(d)-1 if len(d) > 2 else 2))
    km = KMeans(n_clusters=k, random_state=42, n_init=10)
    km.fit(Xs)
    km_dist = km.transform(Xs).min(axis=1)
    km_cut = np.percentile(km_dist, max(80, int((1-contamination)*100))) if len(km_dist) > 5 else km_dist.mean()
    km_pred = (km_dist >= km_cut).astype(int)

    zmax = np.abs(Xs).max(axis=1) if len(d) > 0 else np.array([])
    z_pred = (zmax > 2.8).astype(int) if len(d) > 0 else np.array([])

    result = d.copy()
    result['ml_iso_flag'] = iso_pred
    result['ml_lof_flag'] = lof_pred
    result['ml_svm_flag'] = svm_pred
    result['ml_kmeans_flag'] = km_pred
    result['ml_zscore_flag'] = z_pred
    result['ml_iso_score'] = iso_score
    result['ml_lof_score'] = lof_score
    result['ml_svm_score'] = svm_score
    result['ml_kmeans_score'] = km_dist
    result['ml_vote_count'] = result[['ml_iso_flag','ml_lof_flag','ml_svm_flag','ml_kmeans_flag','ml_zscore_flag']].sum(axis=1)
    result['ml_anomaly_flag'] = (result['ml_vote_count'] >= 2).astype(int)
    result['ml_risk_level'] = pd.cut(result['ml_vote_count'], bins=[-0.1,1.5,2.5,5], labels=['🟢 Бага','🟠 Өндөр','🔴 Маш өндөр']).astype(str)

    xai_importance = pd.DataFrame({
        'feature': feat_cols,
        'importance': np.abs(np.corrcoef(np.column_stack([Xs, result['ml_anomaly_flag'].values]).T)[-1,:-1]) if len(d) > 3 else np.zeros(len(feat_cols))
    })
    xai_importance['importance'] = xai_importance['importance'].replace([np.inf,-np.inf],0).fillna(0)
    xai_importance = xai_importance.sort_values('importance', ascending=False)

    if shap is not None and len(d) > 20 and result['ml_anomaly_flag'].nunique() > 1:
        try:
            y_ml = result['ml_anomaly_flag'].astype(int).values
            rf_xai = RandomForestClassifier(n_estimators=120, random_state=42, class_weight='balanced')
            rf_xai.fit(X, y_ml)
            explainer = shap.TreeExplainer(rf_xai)
            shap_vals = explainer.shap_values(X)
            if isinstance(shap_vals, list):
                sv = shap_vals[-1]
            else:
                sv = shap_vals
            shap_abs = np.abs(sv).mean(axis=0)
            xai_importance = pd.DataFrame({'feature': feat_cols, 'importance': shap_abs}).sort_values('importance', ascending=False)
            result['xai_top_feature'] = np.array(feat_cols)[np.argmax(np.abs(sv), axis=1)]
        except Exception:
            result['xai_top_feature'] = xai_importance.iloc[0]['feature'] if not xai_importance.empty else ''
    else:
        if not xai_importance.empty:
            topf = xai_importance.iloc[0]['feature']
            result['xai_top_feature'] = np.where(result[topf].fillna(0) != 0, topf, '')
        else:
            result['xai_top_feature'] = ''

    model_summary = pd.DataFrame([
        {'Алгоритм':'Isolation Forest','Илрүүлсэн аномали':int(result['ml_iso_flag'].sum()),'Хувь':round(float(result['ml_iso_flag'].mean()*100),2)},
        {'Алгоритм':'Local Outlier Factor','Илрүүлсэн аномали':int(result['ml_lof_flag'].sum()),'Хувь':round(float(result['ml_lof_flag'].mean()*100),2)},
        {'Алгоритм':'One-Class SVM','Илрүүлсэн аномали':int(result['ml_svm_flag'].sum()),'Хувь':round(float(result['ml_svm_flag'].mean()*100),2)},
        {'Алгоритм':'KMeans distance','Илрүүлсэн аномали':int(result['ml_kmeans_flag'].sum()),'Хувь':round(float(result['ml_kmeans_flag'].mean()*100),2)},
        {'Алгоритм':'Z-score','Илрүүлсэн аномали':int(result['ml_zscore_flag'].sum()),'Хувь':round(float(result['ml_zscore_flag'].mean()*100),2)},
        {'Алгоритм':'Ensemble ≥2 votes','Илрүүлсэн аномали':int(result['ml_anomaly_flag'].sum()),'Хувь':round(float(result['ml_anomaly_flag'].mean()*100),2)},
    ])
    return result, feat_cols, model_summary, xai_importance


def render_xai_summary(xai_importance, top_n=10):
    if xai_importance is None or len(xai_importance) == 0:
        st.info('XAI тайлбар гаргах хангалттай өгөгдөл алга.')
        return
    st.markdown('#### 🔎 XAI — Эрсдэлийг хамгийн их тайлбарлаж буй шинжүүд')
    show = xai_importance.head(top_n).copy()
    show['importance'] = pd.to_numeric(show['importance'], errors='coerce').fillna(0)
    st.dataframe(show, use_container_width=True, hide_index=True)
    fig_xai = px.bar(show.sort_values('importance', ascending=True), x='importance', y='feature', orientation='h', title='XAI нөлөөллийн эрэмбэ')
    st.plotly_chart(fig_xai, use_container_width=True)

def run_ml(tb_all, cont, n_est):
    df = tb_all.copy()
    if df.empty or len(df) < 5:
        return pd.DataFrame(), np.array([]), np.array([]), [], {}, '', pd.DataFrame(), np.array([])
    needed = ['account_code','turnover_debit','turnover_credit','closing_debit','closing_credit','opening_debit','net_change_signed','year']
    for c in needed:
        if c not in df.columns:
            df[c] = 0 if c != 'account_code' else ''
    df['cat_code'] = df['account_code'].astype(str).str[:3]
    le = LabelEncoder()
    df['cat_num'] = le.fit_transform(df['cat_code'].fillna(''))
    df['log_turn_d'] = np.log1p(pd.to_numeric(df['turnover_debit'], errors='coerce').fillna(0).abs())
    df['log_turn_c'] = np.log1p(pd.to_numeric(df['turnover_credit'], errors='coerce').fillna(0).abs())
    df['log_close_d'] = np.log1p(pd.to_numeric(df['closing_debit'], errors='coerce').fillna(0).abs())
    df['log_close_c'] = np.log1p(pd.to_numeric(df['closing_credit'], errors='coerce').fillna(0).abs())
    df['turn_ratio'] = (pd.to_numeric(df['turnover_debit'], errors='coerce').fillna(0) / pd.to_numeric(df['turnover_credit'], errors='coerce').replace(0, np.nan)).fillna(0).replace([np.inf, -np.inf], 0)
    if 'net_change_signed' in df.columns:
        df['log_abs_change'] = np.log1p(pd.to_numeric(df['net_change_signed'], errors='coerce').fillna(0).abs())
    else:
        df['log_abs_change'] = np.log1p((pd.to_numeric(df['closing_debit'], errors='coerce').fillna(0) - pd.to_numeric(df['opening_debit'], errors='coerce').fillna(0)).abs())
    # ISA 520: Аналитик горим — өсөлтийн хурд нэмэх
    opening_bal = pd.to_numeric(df.get('opening_debit', 0), errors='coerce').fillna(0).abs() + pd.to_numeric(df.get('opening_credit', 0), errors='coerce').fillna(0).abs()
    closing_bal = pd.to_numeric(df.get('closing_debit', 0), errors='coerce').fillna(0).abs() + pd.to_numeric(df.get('closing_credit', 0), errors='coerce').fillna(0).abs()
    df['growth_rate'] = np.where(opening_bal > 0, (closing_bal - opening_bal) / opening_bal, 0)
    df['growth_rate'] = df['growth_rate'].clip(-10, 10).fillna(0)
    feats = ['cat_num', 'log_turn_d', 'log_turn_c', 'log_close_d', 'log_close_c', 'turn_ratio', 'log_abs_change', 'growth_rate', 'year']
    X = df[feats].fillna(0).replace([np.inf, -np.inf], 0)
    iso = IsolationForest(contamination=min(max(cont, 0.01), 0.4), random_state=42, n_estimators=200)
    df['iso_anomaly'] = (iso.fit_predict(X) == -1).astype(int)
    sc = StandardScaler()
    df['zscore_anomaly'] = (np.abs(sc.fit_transform(X)).max(axis=1) > 2.0).astype(int)
    p95 = df['turn_ratio'].quantile(0.95)
    df['turn_anomaly'] = ((df['turn_ratio'] > p95) | (df['turn_ratio'] < -p95)).astype(int)
    df['ensemble_anomaly'] = ((df['iso_anomaly'] == 1) | ((df['zscore_anomaly'] == 1) & (df['turn_anomaly'] == 1))).astype(int)
    y = df['ensemble_anomaly'].values
    if len(np.unique(y)) < 2 or len(df) < 10:
        fi = pd.DataFrame({'feature': feats, 'importance': [0.0]*len(feats)})
        res = {'Random Forest': {'pred': y, 'prob': np.zeros(len(y)), 'precision': 0.0, 'recall': 0.0, 'f1': 0.0, 'auc': 0.0}}
        return df, X, y, feats, res, 'Random Forest', fi, np.zeros(len(df), dtype=int)
    n_splits = min(5, int(np.bincount(y).min())) if np.bincount(y).min() > 1 else 2
    cv = StratifiedKFold(n_splits=max(2, n_splits), shuffle=True, random_state=42)
    models = {
        'Random Forest': RandomForestClassifier(n_estimators=n_est, max_depth=10, random_state=42, class_weight='balanced'),
        'Gradient Boosting': GradientBoostingClassifier(n_estimators=150, learning_rate=0.1, random_state=42),
        'Logistic Regression': LogisticRegression(max_iter=1000, random_state=42, class_weight='balanced'),
    }
    res = {}
    for nm, mdl in models.items():
        try:
            yp = cross_val_predict(mdl, X, y, cv=cv, method='predict')
            ypr = cross_val_predict(mdl, X, y, cv=cv, method='predict_proba')[:, 1]
            res[nm] = {'pred': yp, 'prob': ypr, 'precision': precision_score(y, yp, zero_division=0), 'recall': recall_score(y, yp, zero_division=0), 'f1': f1_score(y, yp, zero_division=0), 'auc': roc_auc_score(y, ypr)}
        except Exception:
            yp = np.zeros(len(y), dtype=int)
            ypr = np.zeros(len(y), dtype=float)
            res[nm] = {'pred': yp, 'prob': ypr, 'precision': 0.0, 'recall': 0.0, 'f1': 0.0, 'auc': 0.0}
    best = max(res, key=lambda k: res[k]['f1']) if res else ''
    rf = models['Random Forest']
    try:
        rf.fit(X, y)
        fi = pd.DataFrame({'feature': feats, 'importance': rf.feature_importances_}).sort_values('importance', ascending=False)
    except Exception:
        fi = pd.DataFrame({'feature': feats, 'importance': [0.0]*len(feats)})
    nt = len(df)
    ns = max(1, int(nt * 0.20))
    at = pd.to_numeric(df['turnover_debit'], errors='coerce').fillna(0).abs() + pd.to_numeric(df['turnover_credit'], errors='coerce').fillna(0).abs()
    wt = (at / at.sum()).fillna(1 / nt) if at.sum() != 0 else pd.Series(np.repeat(1/nt, nt))
    np.random.seed(42)
    ms = np.zeros(nt, dtype=int)
    ms[np.random.choice(nt, size=min(ns, nt), replace=False, p=wt.values)] = 1
    ym = (ms & y).astype(int)
    return df, X, y, feats, res, best, fi, ym

# ═══════════════════════════════════════
# 🏷️ ДАНСНЫ АНГИЛАЛ — ХАСАХ БҮЛГҮҮД
# ═══════════════════════════════════════
# 6 бүлэг: шимтгэл, хаалтын бичилт, идэвхгүй, тогтмол зардал, коммунал, үндсэн орлого

EXCL_RULES = {
    'nan_data': {
        'label': '⚠️ NaN / дутуу өгөгдөл',
        'help': 'account_code, account_name, description, counterparty дутуу мөрүүд',
        'default': True,
        'account_prefixes': [],
        'name_keywords': ['nan','none','unknown'],
        'desc_keywords': ['nan','none','unknown'],
    },
    'шимтгэл': {
        'label': '🏦 Шимтгэл, хураамж, татвар',
        'help': 'Банкны шимтгэл, ХХОАт, НДШ, НӨАт, хүү зэрэг давтамжтай, бага дүнтэй бичилтүүд',
        'default': True,
        'account_prefixes': ['7027','7028','7029','354','356','3541','3542','3543','3544'],
        'name_keywords': [
            'шимтгэл','хураамж','банкны шимтгэл','үйлчилгээний хураамж',
            'комисс','commission','fee','bank charge','service charge',
            'тэмдэгтийн хураамж','нийгмийн даатгал','ндш','ххоат','нөат','vat',
            'хүү','interest','алданги','торгууль','penalty',
        ],
        'desc_keywords': [
            'шимтгэл','хураамж','commission','fee','interest','хүү','алданги',
        ],
    },
    'хаалтын_бичилт': {
        'label': '📕 Хаалтын бичилт, залруулга',
        'help': 'Жилийн эцсийн хаалт, залруулга, буцаалт, сторно, нээлтийн бичилтүүд',
        'default': True,
        'account_prefixes': [],
        'name_keywords': [],
        'desc_keywords': [
            'хаалт','хаах','closing','close','year end','year-end',
            'жилийн эцсийн','хаалтын бичилт','тайлант үеийн хаалт',
            'залруулга','adjustment','adjusting','аудитын залруулга',
            'буцаалт','reversal','сторно','storno',
            'нээлтийн бичилт','opening entry','нээлт',
            'хуримтлагдсан элэгдэл','элэгдэл тооцох','depreciation',
        ],
    },
    'идэвхгүй': {
        'label': '⏸️ Идэвхгүй данс (эргэлтгүй)',
        'help': 'Тухайн жилд ямар ч эргэлтгүй (дебит=0, кредит=0) данснууд',
        'default': True,
        'account_prefixes': [],
        'name_keywords': [],
        'desc_keywords': [],
    },
    'тогтмол_зардал': {
        'label': '📋 Тогтмол зардал (цалин, түрээс г.м.)',
        'help': 'Цалин, НДШ, түрээс, даатгал, элэгдэл зэрэг сар бүр давтагддаг зардлууд',
        'default': False,
        'account_prefixes': ['701','702','703','704','706','710','711','712','713','714','7011','7012','7013','7014','7021','7022','7023'],
        'name_keywords': [
            'цалин','хөдөлмөрийн хөлс','salary','wage','цалингийн',
            'түрээс','rent','lease','түрээсийн',
            'даатгал','insurance','даатгалын',
            'элэгдэл','depreciation','хорогдол','amortization',
            'нөөц','provision','нөөцийн',
            'тэтгэмж','тэтгэвэр','pension',
            'урамшуулал','bonus',
        ],
        'desc_keywords': [
            'цалин','salary','түрээс','rent','даатгал','insurance',
            'элэгдэл','depreciation','нөөц','provision',
        ],
    },
    'коммунал': {
        'label': '💡 Коммунал (тог, ус, дулаан, холбоо)',
        'help': 'Цахилгаан, ус, дулаан, интернет, утас, шуудан зэрэг коммунал зардлууд',
        'default': False,
        'account_prefixes': ['7024','7025','7026'],
        'name_keywords': [
            'цахилгаан','электр','electricity','power',
            'ус','усны','water',
            'дулаан','дулааны','heating','heat',
            'тог','тогны',
            'холбоо','холбооны','утас','утасны','telephone','phone','telecom',
            'интернет','internet','сүлжээ','network',
            'шуудан','шуудангийн','postal',
            'коммунал','utility','utilities',
        ],
        'desc_keywords': [
            'цахилгаан','electricity','ус','water','дулаан','heating',
            'тог','утас','phone','интернет','internet','коммунал','utility',
        ],
    },
    'үндсэн_орлого': {
        'label': '💰 Үндсэн үйл ажиллагааны орлого',
        'help': 'Борлуулалтын орлого, үйлчилгээний орлого — бизнесийн үндсэн урсгал',
        'default': False,
        'account_prefixes': ['511','512','521','522','531','532','601','602','611','612'],
        'name_keywords': [
            'борлуулалтын орлого','борлуулалт','sales revenue','revenue',
            'үйлчилгээний орлого','service revenue','service income',
            'үндсэн үйл ажиллагааны орлого','operating revenue',
            'бараа борлуулсны орлого','бүтээгдэхүүн борлуулалт',
            'ажил үйлчилгээний орлого',
        ],
        'desc_keywords': [
            'борлуулалт','sales','орлого','revenue','income',
        ],
    },
}

def classify_exclusions(df, level='account'):
    """Данс/гүйлгээг 6 хасах ангилалд хуваана.
    Returns: df with 'exclusion_tag' column
    Tags: 'шимтгэл','хаалтын_бичилт','идэвхгүй','тогтмол_зардал','коммунал','үндсэн_орлого','' (хасахгүй)
    """
    d = df.copy()
    d['exclusion_tag'] = ''
    code_str = d['account_code'].astype(str) if 'account_code' in d.columns else pd.Series('', index=d.index)
    name_lower = d['account_name'].astype(str).str.lower() if 'account_name' in d.columns else pd.Series('', index=d.index)

    if level == 'transaction':
        desc_lower = d['transaction_description'].astype(str).str.lower() if 'transaction_description' in d.columns else pd.Series('', index=d.index)
        combined = name_lower + ' ' + desc_lower
    else:
        combined = name_lower

    # ── NaN / дутуу өгөгдөл ──
    if level == 'transaction':
        has_missing = (code_str.str.strip() == '') | (combined.str.strip() == '')
        if 'counterparty_name' in d.columns:
            has_missing = has_missing | (d['counterparty_name'].astype(str).str.strip() == '')
        d.loc[has_missing, 'exclusion_tag'] = 'nan_data'
    else:
        has_missing = (code_str.str.strip() == '') | (name_lower.str.strip() == '')
        d.loc[has_missing, 'exclusion_tag'] = 'nan_data'

    # ── Дүрмүүдийг дарааллаар хэрэглэх (эхнийх нь давуу) ──
    for tag, rule in EXCL_RULES.items():
        if tag in ('идэвхгүй','nan_data'):
            continue  # Тусгай шалгалтууд
        untagged = d['exclusion_tag'] == ''
        # Дансны код prefix
        for prefix in rule.get('account_prefixes', []):
            mask = code_str.str.startswith(prefix) & untagged
            d.loc[mask, 'exclusion_tag'] = tag
        # Нэр/тайлбар keyword
        kws = rule.get('desc_keywords', []) if level == 'transaction' else rule.get('name_keywords', [])
        for kw in kws:
            mask = combined.str.contains(kw, na=False, regex=False) & (d['exclusion_tag'] == '')
            d.loc[mask, 'exclusion_tag'] = tag

    # ── Идэвхгүй данс (эргэлт = 0) — зөвхөн дансны түвшинд ──
    if level == 'account':
        for c in ['turnover_debit', 'turnover_credit']:
            if c not in d.columns: d[c] = 0
        turn_total = pd.to_numeric(d['turnover_debit'], errors='coerce').fillna(0).abs() + \
                     pd.to_numeric(d['turnover_credit'], errors='coerce').fillna(0).abs()
        d.loc[(turn_total == 0) & (d['exclusion_tag'] == ''), 'exclusion_tag'] = 'идэвхгүй'

    return d


# ═══════════════════════════════════════
# 🧠 УХААЛАГ ФАЙЛ ТАНИХ СИСТЕМ
# ═══════════════════════════════════════
def detect_file_type(f):
    """Файлын төрлийг автоматаар таних. Returns: (type, year)
    Types: 'raw_tb', 'edt', 'tb_std', 'ledger', 'part1', 'unknown'
    """
    name = f.name.lower()
    fname_orig = f.name
    year = get_year(f.name)

    # CSV/GZ → Ledger
    if name.endswith('.csv') or name.endswith('.gz') or name.endswith('.csv.gz'):
        return 'ledger', year

    # XLSX → need to check
    if not name.endswith('.xlsx'):
        return 'unknown', year

    # ── Файлын нэрээр хурдан таних ──
    name_check = fname_orig.lower().replace('_', ' ').replace('-', ' ')
    # ЕЖ / Ерөнхий журнал / Journal
    edt_keywords = ['ерөнхий журнал', 'ерөнхий дэвтэр', 'едт', 'edt', 'general ledger', 'general journal',
                    'еренхий журнал', 'journal', 'journal entry', 'journal entries']
    for kw in edt_keywords:
        if kw in name_check:
            return 'edt', year
    # ГҮЙЛГЭЭ_БАЛАНС / Trial Balance / Journal TB
    tb_keywords = ['гүйлгээ баланс', 'гүйлгээ_баланс', 'гуйлгээ баланс', 'trial balance',
                   'гүйлгэ баланс', 'гуйлгэ баланс']
    for kw in tb_keywords:
        if kw in name_check:
            return 'raw_tb', year
    # TB_standardized
    if 'tb_standardized' in name_check or 'tb standardized' in name_check:
        return 'tb_std', year
    # Part1
    if 'part1' in name_check or 'part 1' in name_check:
        return 'part1', year
    # Ledger
    if 'ledger' in name_check or 'prototype_ledger' in name_check:
        return 'ledger', year

    # ── Sheet бүтцээр таних ──
    import openpyxl
    try:
        raw = f.read()
        f.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        sheets = wb.sheetnames

        # TB_standardized: has '02_ACCOUNT_SUMMARY' sheet
        if '02_ACCOUNT_SUMMARY' in sheets:
            if '04_RISK_MATRIX' in sheets:
                wb.close()
                return 'part1', year
            wb.close()
            return 'tb_std', year

        # Part1: has '04_RISK_MATRIX' sheet
        if '04_RISK_MATRIX' in sheets:
            wb.close()
            return 'part1', year

        # ── Агуулгаар таних (200 мөр хүртэл шалгана) ──
        ws = wb[sheets[0]]
        sample_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            sample_rows.append(row)
            if i >= 200:
                break
        wb.close()

        # ЕЖ: contains "Данс:" or "Компани:" or "ЕРӨНХИЙ" or "Журнал:" pattern
        for row in sample_rows:
            if row[0] is not None:
                s = str(row[0]).strip()
                if s.startswith('Данс:') or s.startswith('Компани:') or s.startswith('ЕРӨНХИЙ') or s.startswith('Журнал:'):
                    return 'edt', year
            # ЕЖ: column with "Данс:" might be in other columns too
            for cell in row[:5]:
                if cell is not None and 'Данс:' in str(cell):
                    return 'edt', year

        # ГҮЙЛГЭЭ_БАЛАНС: has account codes like 101-XX-XX-XXX in column B
        for row in sample_rows:
            if len(row) >= 2 and row[1] is not None:
                code = str(row[1]).strip()
                if re.match(r'\d{3}-\d{2}-\d{2}-\d{3}', code):
                    return 'raw_tb', year

        # Fallback: check if it looks like a balance sheet
        for row in sample_rows:
            if row[0] is not None:
                try:
                    int(float(row[0]))
                    if len(row) >= 8 and row[1] is not None and re.match(r'\d{3}-', str(row[1])):
                        return 'raw_tb', year
                except:
                    pass

        return 'unknown', year
    except Exception:
        f.seek(0)
        return 'unknown', year


def parse_account_names(file_obj):
    """Дансны код + нэрийн лавлах файл уншина.
    Формат: A баганад дансны код (1, 31, 312, 3121, 31213, ...), B баганад нэр.
    Санхүүгийн байдлын тайлан (СТ-1А) эсвэл дансны жагсаалт файл дэмжинэ.
    """
    import openpyxl
    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        code_map = {}
        for row in ws.iter_rows(values_only=True):
            c0 = str(row[0]).strip() if row[0] is not None else ''
            c1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            if not c0 or not c1:
                continue
            # Зөвхөн тоон код авах (1, 31, 312, 3121, 31213, ...)
            c0_clean = re.sub(r'[^0-9]', '', c0)
            if c0_clean and len(c0_clean) >= 1:
                code_map[c0_clean] = c1.strip()
        wb.close()
        return code_map
    except Exception:
        return {}

def merge_account_names(df, code_map):
    """Гүйлгээний DataFrame-д дансны нэрийг prefix matching-аар нэгтгэнэ.
    Жишээ: 312130201 → 31213 → 'Арилжааны банк дахь харилцах'
    """
    if not code_map or 'account_code' not in df.columns:
        return df
    d = df.copy()

    def _find_name(code):
        code_str = re.sub(r'[^0-9]', '', str(code))
        # Урт prefix-ээс богино руу хайна (хамгийн нарийвчлалтай нэрийг олно)
        for length in range(len(code_str), 0, -1):
            prefix = code_str[:length]
            if prefix in code_map:
                return code_map[prefix]
        return ''

    # Хоосон эсвэл байхгүй нэртэй мөрүүдэд нэр нэмэх
    if 'account_name' not in d.columns:
        d['account_name'] = ''
    mask = d['account_name'].fillna('').str.strip() == ''
    if mask.any():
        d.loc[mask, 'account_name'] = d.loc[mask, 'account_code'].apply(_find_name)

    return d

def detect_account_names_file(file_obj):
    """Дансны нэрийн лавлах файл мөн эсэхийг шалгана."""
    import openpyxl
    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        score = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 20: break
            row_text = ' '.join(str(c).strip().lower() for c in row if c is not None)
            if 'дансны код' in row_text: score += 3
            if 'балансын үзүүлэлт' in row_text or 'дансны нэр' in row_text: score += 3
            if 'санхүүгийн байдл' in row_text or 'ст-1' in row_text: score += 2
            if 'эхний үлдэгдэл' in row_text or 'эцсийн үлдэгдэл' in row_text: score += 1
        wb.close()
        file_obj.seek(0)
        return score >= 3
    except:
        file_obj.seek(0)
        return False

def materiality_base_from_tb(tb_df):
    if tb_df is None or tb_df.empty:
        return 0.0
    candidates = []
    for c in ['turnover_debit','turnover_credit','closing_debit','closing_credit','opening_debit','opening_credit']:
        if c in tb_df.columns:
            candidates.append(pd.to_numeric(tb_df[c], errors='coerce').fillna(0).abs().sum())
    return float(max(candidates)) if candidates else 0.0

def build_materiality_by_account(tb_df, overall_materiality, performance_ratio=0.75, trivial_ratio=0.05):
    """ISA 320 + ISA 330: Данс тус бүрийн материаллаг байдал + аудитын горим.
    Эрсдэлийн коэффициент (ISA 320.A12): Дансны өөрчлөлт, ангиллаас хамааран залруулна.
    Аудитын горим (ISA 330): Эрсдэлийн түвшнээс хамааран горим санал болгоно.
    """
    if tb_df is None or tb_df.empty:
        return pd.DataFrame()
    d = tb_df.copy()
    for c in ['account_code','account_name','closing_debit','closing_credit',
              'turnover_debit','turnover_credit','opening_debit','opening_credit']:
        if c not in d.columns:
            d[c] = '' if c in ['account_code','account_name'] else 0.0
    for c in ['closing_debit','closing_credit','turnover_debit','turnover_credit','opening_debit','opening_credit']:
        d[c] = pd.to_numeric(d[c], errors='coerce').fillna(0)

    # ── Суурь дүн тооцох ──
    d['closing_abs'] = d['closing_debit'].abs() + d['closing_credit'].abs()
    d['turnover_abs'] = d['turnover_debit'].abs() + d['turnover_credit'].abs()
    d['суурь_дүн'] = np.where(d['closing_abs'] > 0, d['closing_abs'], d['turnover_abs'])

    # ── Дансны ангилал (ISA 315) ──
    d['ангилал'] = d['account_code'].astype(str).str[0].map(
        {'1':'Хөрөнгө','2':'Өр төлбөр','3':'Эздийн өмч',
         '4':'Зардал','5':'Орлого','6':'Орлого',
         '7':'Үйл ажиллагааны зардал','8':'Бусад зардал','9':'Нэгдсэн данс'}
    ).fillna('Бусад')

    # ── Аналитик горим: Өөрчлөлтийн хувь (ISA 520) ──
    opening = d['opening_debit'].abs() + d['opening_credit'].abs()
    d['өөрчлөлт_%'] = np.where(opening > 0, (d['суурь_дүн'] - opening) / opening * 100, 0).round(1)

    # ── Эрсдэлийн коэффициент (ISA 320.A12) ──
    def _risk_coeff(row):
        pct = abs(row.get('өөрчлөлт_%', 0))
        cat = str(row.get('account_code', ''))[:1]
        if pct > 50: return 0.50   # Маш өндөр өөрчлөлт → 2× бага босго
        if pct > 30 and cat == '1': return 0.60  # Хөрөнгийн өндөр өөрчлөлт
        if cat in ('5','6','7','8'): return 0.75  # Орлого/зардал субъектив
        if pct > 20: return 0.75
        if pct < 5: return 1.20   # Бага эрсдэл → илүү өндөр босго
        return 1.00
    d['эрсдэлийн_коэфф'] = d.apply(_risk_coeff, axis=1)

    # ── Материаллаг байдлын хуваарилалт ──
    total_base = max(d['суурь_дүн'].sum(), 1)
    d['жин_%'] = (d['суурь_дүн'] / total_base * 100).round(3)
    d['зөвшөөрөгдөх_алдаа'] = (d['жин_%'] / 100 * overall_materiality * d['эрсдэлийн_коэфф']).round(0)
    d['гүйцэтгэлийн_мат'] = (d['зөвшөөрөгдөх_алдаа'] * performance_ratio).round(0)
    d['анхаарах_доод'] = (d['зөвшөөрөгдөх_алдаа'] * trivial_ratio).round(0)

    # ── Босго давсан эсэх (ISA 320.A12) ──
    d['босго_давсан'] = np.where(d['суурь_дүн'] > d['зөвшөөрөгдөх_алдаа'], '⚠️ Тийм', '✅ Үгүй')

    # ── Эрсдэлийн түвшин ──
    risk_score = d['жин_%'] * (2 - d['эрсдэлийн_коэфф'])
    d['эрсдэлийн_түвшин'] = pd.cut(risk_score, bins=[-0.001, 1.0, 5.0, 100.0],
        labels=['Бага', 'Дунд', 'Өндөр']).astype(str)

    # ── ISA 330 аудитын горимын санал ──
    def _audit_proc(row):
        if row.get('босго_давсан') == '⚠️ Тийм':
            return 'Нарийвчилсан шалгалт + Баталгаажуулалт (ISA 505)'
        lv = row.get('эрсдэлийн_түвшин', 'Бага')
        if lv == 'Өндөр': return 'Нарийвчилсан шалгалт (ISA 330.18)'
        if lv == 'Дунд': return 'Шинжилгээний процедур (ISA 520) + Хязгаарлагдмал шалгалт'
        return 'Шинжилгээний процедур (ISA 520)'
    d['аудитын_горим'] = d.apply(_audit_proc, axis=1)

    out = d[['account_code','account_name','ангилал','суурь_дүн','turnover_abs',
        'өөрчлөлт_%','жин_%','эрсдэлийн_коэфф',
        'зөвшөөрөгдөх_алдаа','гүйцэтгэлийн_мат','анхаарах_доод',
        'босго_давсан','эрсдэлийн_түвшин','аудитын_горим']].copy()
    out.columns = ['Дансны код','Дансны нэр','Ангилал','Эцсийн үлдэгдэл','Нийт эргэлт',
        'Өөрчлөлт %','Жин %','Эрсдэлийн коэфф',
        'Зөвшөөрөгдөх алдаа ₮','Гүйцэтгэлийн мат ₮','Анхаарах доод ₮',
        'Босго давсан','Эрсдэлийн түвшин','Аудитын горим (ISA 330)']
    return out.sort_values('Зөвшөөрөгдөх алдаа ₮', ascending=False).reset_index(drop=True)



FILE_TYPE_LABELS = {
    'raw_tb': ('📗 ГҮЙЛГЭЭ_БАЛАНС', 'Гүйлгээ-балансын түүхий файл → TB болгон хөрвүүлнэ'),
    'edt': ('📘 Ерөнхий журнал (ЕЖ)', 'Ерөнхий журналын гүйлгээ → стандарт CSV ledger болгон хөрвүүлнэ'),
    'tb_std': ('📊 TB_standardized', 'Стандартчилсан TB → шинжилгээнд бэлэн'),
    'ledger': ('📄 Ledger / CSV', 'Гүйлгээний дэлгэрэнгүй файл → journal шинжилгээнд бэлэн'),
    'part1': ('📈 Part1', 'Сарын нэгтгэл + Эрсдэлийн матриц → TB шинжилгээнд бэлэн'),
    'unknown': ('❓ Тодорхойгүй', 'Файлын төрлийг таних боломжгүй'),
}

for _k, _v in {
    'prepared_tb_cache': {},
    'prepared_part1_cache': {},
    'prepared_ledger_cache': {},
    'prep_detected_rows': [],
    'tb_error': '',
    'journal_error': '',
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


def _cache_add(cache_key, filename, raw_bytes):
    st.session_state[cache_key][filename] = raw_bytes


def _cache_files(cache_key):
    out = []
    for name, raw in st.session_state.get(cache_key, {}).items():
        bio = io.BytesIO(raw)
        bio.name = name
        out.append(bio)
    return out


def _df_to_csv_bytes(df):
    return df.to_csv(index=False).encode('utf-8-sig')


def _df_to_excel_bytes(df_map):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        for sname, df in df_map.items():
            df.to_excel(w, sheet_name=sname[:31], index=False)
    return buf.getvalue()


def _prepare_from_uploaded(uploaded, acct_name_map=None):
    detected_rows = []
    acct_name_map = acct_name_map or {}
    for f in uploaded or []:
        try:
            ftype, year = detect_file_type(f)
            f.seek(0)
            label, desc = FILE_TYPE_LABELS.get(ftype, FILE_TYPE_LABELS['unknown'])
            detected_rows.append({'Файл': f.name, 'Төрөл': label, 'Он': year, 'Тайлбар': desc})
            if ftype == 'raw_tb':
                buf, tb_sum = process_raw_tb(f)
                if tb_sum is not None and not tb_sum.empty:
                    _cache_add('prepared_tb_cache', f'TB_standardized_{year}.xlsx', buf.getvalue())
            elif ftype == 'tb_std':
                _cache_add('prepared_tb_cache', f.name, f.getvalue())
            elif ftype == 'part1':
                _cache_add('prepared_part1_cache', f.name, f.getvalue())
            elif ftype == 'ledger':
                _cache_add('prepared_ledger_cache', f.name, f.getvalue())
                try:
                    f.seek(0)
                    led_df = read_ledger(f)
                    if not led_df.empty:
                        _, _, _, rm, _ = generate_part1(led_df, year)
                        mo = pd.read_excel(io.BytesIO(_df_to_excel_bytes({'tmp': rm})), sheet_name='tmp') if False else None
                except Exception:
                    pass
            elif ftype == 'edt':
                f.seek(0)
                edt_df, cnt = process_edt(f, year)
                if acct_name_map and not edt_df.empty:
                    edt_df = merge_account_names(edt_df, acct_name_map)
                if cnt > 0 and not edt_df.empty:
                    csv_name = f'ledger_from_EJ_{year}_{Path(f.name).stem}.csv'
                    _cache_add('prepared_ledger_cache', csv_name, _df_to_csv_bytes(edt_df))
        except Exception as e:
            detected_rows.append({'Файл': f.name, 'Төрөл': '❌ Алдаа', 'Он': '', 'Тайлбар': str(e)})
    st.session_state['prep_detected_rows'] = detected_rows


def _build_part1_from_prepared_ledgers():
    ledger_files = _cache_files('prepared_ledger_cache')
    created = 0
    for lf in ledger_files:
        try:
            year = get_year(lf.name)
            lf.seek(0)
            led_df = read_ledger(lf)
            if led_df.empty:
                continue
            part1_buf, _, _, _, _ = generate_part1(led_df, year)
            out_name = f'Part1_generated_{year}_{Path(lf.name).stem[:40]}.xlsx'
            _cache_add('prepared_part1_cache', out_name, part1_buf.getvalue())
            created += 1
        except Exception:
            continue
    return created


def _render_downloads(title, cache_key, mime):
    files = st.session_state.get(cache_key, {})
    if files:
        st.markdown(title)
        for name, raw in files.items():
            st.download_button(f'📥 {name}', raw, file_name=name, mime=mime, key=f'dl_{cache_key}_{name}')


def _show_dataframe_download(df, filename, label='📥 CSV татах'):
    if df is not None and not df.empty:
        st.download_button(label, _df_to_csv_bytes(df), file_name=filename, mime='text/csv', key=f'dl_{filename}')


if page.startswith("1"):
    st.header("1️⃣ Өгөгдөл оруулах, бэлтгэх")
    st.markdown("Файлаа нэг удаа оруулаад дараагийн цэсүүд дээр дахин ашиглаж болно.")

    uploaded = st.file_uploader("📎 Бүх файлуудаа энд оруулна уу", type=['xlsx', 'csv', 'gz'], accept_multiple_files=True, key='smart_prep_main')
    acct_name_file = st.file_uploader("📋 Дансны нэрийн лавлах файл (заавал биш)", type=['xlsx'], key='acct_names_prep_main')
    acct_name_map = parse_account_names(acct_name_file) if acct_name_file else {}

    cpa, cpb = st.columns([1,1])
    with cpa:
        if st.button('🛠️ Файлуудыг таньж бэлтгэх', type='primary', use_container_width=True):
            if uploaded:
                _prepare_from_uploaded(uploaded, acct_name_map)
                created = _build_part1_from_prepared_ledgers()
                st.success(f'✅ Файлууд бэлтгэгдлээ. Ledger-ээс {created} Part1 файл үүсгэлээ.')
            else:
                st.warning('Файл оруулна уу.')
    with cpb:
        if st.button('🧹 Бүх хадгалсан өгөгдлийг цэвэрлэх', use_container_width=True):
            for key in ['prepared_tb_cache','prepared_part1_cache','prepared_ledger_cache','prep_detected_rows',
                        'tb_analysis_done','journal_ai_done','tb_all','tb_stats','rm_all','mo_all','tb_filtered',
                        'tb_ml_df','tb_feature_importance','journal_ml_result','journal_ml_show',
                        'journal_model_summary','journal_xai','journal_ledger_stats','tb_upload_cache','journal_upload_cache']:
                if key in st.session_state:
                    if isinstance(st.session_state[key], dict): st.session_state[key] = {}
                    elif isinstance(st.session_state[key], list): st.session_state[key] = []
                    elif isinstance(st.session_state[key], bool): st.session_state[key] = False
                    else: st.session_state[key] = pd.DataFrame() if 'df' in key or key in ['tb_all','rm_all','mo_all','tb_filtered','tb_ml_df','tb_feature_importance','journal_ml_result','journal_ml_show','journal_model_summary','journal_xai'] else ''
            st.success('Session цэвэрлэгдлээ.')

    if st.session_state.get('prep_detected_rows'):
        st.dataframe(pd.DataFrame(st.session_state['prep_detected_rows']), use_container_width=True, hide_index=True)

    ca, cb, cc = st.columns(3)
    with ca:
        st.metric('TB файл', len(st.session_state.get('prepared_tb_cache', {})))
    with cb:
        st.metric('Ledger файл', len(st.session_state.get('prepared_ledger_cache', {})))
    with cc:
        st.metric('Part1 файл', len(st.session_state.get('prepared_part1_cache', {})))

    _render_downloads('### 📦 Бэлэн болсон TB файлууд', 'prepared_tb_cache', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    _render_downloads('### 📦 Бэлэн болсон Ledger файлууд', 'prepared_ledger_cache', 'text/csv')
    _render_downloads('### 📦 Бэлэн болсон Part1 файлууд', 'prepared_part1_cache', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

elif page.startswith("2"):
    st.header("2️⃣ Гүйлгээ балансын шинжилгээ")
    st.markdown("Энэ хэсэг зөвхөн TB болон Part1-ээр ажиллана. Ерөнхий журналын ML шинжилгээг 3️⃣ цэсэнд хийнэ.")

    all_files = st.file_uploader("📎 Нэмэлт TB / Part1 файл оруулах", type=['xlsx'], accept_multiple_files=True, key='tb_page_upload')
    if all_files:
        st.session_state.tb_upload_cache = uploaded_files_to_cache(all_files)
    elif st.session_state.get('tb_upload_cache'):
        all_files = cache_to_file_objects(st.session_state.tb_upload_cache)
        st.info(f"💾 Өмнө оруулсан {len(all_files)} TB/Part1 файл хадгалагдсан байна.")

    tb_files = _cache_files('prepared_tb_cache')
    p1_files = _cache_files('prepared_part1_cache')
    detect_rows = []
    for f in all_files or []:
        ftype, year = detect_file_type(f)
        f.seek(0)
        detect_rows.append({'Файл': f.name, 'Төрөл': FILE_TYPE_LABELS.get(ftype, FILE_TYPE_LABELS['unknown'])[0], 'Он': year})
        if ftype == 'raw_tb':
            buf, tb_sum = process_raw_tb(f)
            if tb_sum is not None and not tb_sum.empty:
                bio = io.BytesIO(buf.getvalue()); bio.name = f'TB_standardized_{year}_{Path(f.name).stem}.xlsx'
                tb_files.append(bio)
        elif ftype == 'tb_std':
            tb_files.append(f)
        elif ftype == 'part1':
            p1_files.append(f)
    if detect_rows:
        st.dataframe(pd.DataFrame(detect_rows), use_container_width=True, hide_index=True)

    c1s, c2s = st.columns(2)
    with c1s:
        cont = st.slider("🎯 Хэвийн бус дансны хувь (Isolation Forest)", 0.05, 0.20, 0.10, 0.01, key='tb_cont_work')
    with c2s:
        nest = st.slider("🌲 Random Forest модны тоо", 50, 500, 200, 50, key='tb_nest_work')

    with st.expander("🏷️ Эрсдэлийн шинжилгээнээс хасах бүлгүүд", expanded=False):
        excl_settings = {}
        for tag, rule in EXCL_RULES.items():
            excl_settings[tag] = st.checkbox(rule['label'], value=rule.get('default', False), help=rule.get('help',''), key=f'tb_excl_{tag}')

    st.caption(f"Бэлэн TB: {len(tb_files)} файл • Бэлэн Part1: {len(p1_files)} файл")
    if st.button('🚀 TB шинжилгээ эхлүүлэх', type='primary', use_container_width=True, key='run_tb_analysis_main'):
        try:
            tb_all, tb_stats = load_tb(tb_files) if tb_files else (pd.DataFrame(), {})
            rm_all, mo_all = load_part1(p1_files) if p1_files else (pd.DataFrame(), pd.DataFrame())
            tb_show = tb_all.copy()
            if not tb_show.empty:
                tb_show = classify_exclusions(tb_show, level='account')
                active_tags = [k for k, v in excl_settings.items() if v]
                if active_tags:
                    tb_show = tb_show[~tb_show['exclusion_tag'].isin(active_tags)].copy()
                ml_df, X, y, feats, res, best, fi, ym = run_ml(tb_show, cont, nest)
            else:
                ml_df, fi = pd.DataFrame(), pd.DataFrame()
            st.session_state['tb_analysis_done'] = True
            st.session_state['tb_all'] = tb_all
            st.session_state['tb_stats'] = tb_stats
            st.session_state['rm_all'] = rm_all
            st.session_state['mo_all'] = mo_all
            st.session_state['tb_filtered'] = tb_show
            st.session_state['tb_ml_df'] = ml_df
            st.session_state['tb_feature_importance'] = fi
            st.session_state['tb_error'] = ''
            st.success('✅ TB шинжилгээ дууслаа.')
        except Exception as e:
            st.session_state['tb_error'] = str(e)
            st.exception(e)

    if st.session_state.get('tb_error'):
        st.error(st.session_state['tb_error'])

    if st.session_state.get('tb_analysis_done', False):
        tb_stats = st.session_state.get('tb_stats', {})
        rm_all = st.session_state.get('rm_all', pd.DataFrame())
        mo_all = st.session_state.get('mo_all', pd.DataFrame())
        ml_df = st.session_state.get('tb_ml_df', pd.DataFrame())
        fi = st.session_state.get('tb_feature_importance', pd.DataFrame())

        if tb_stats:
            rows = [{'Он': yr, 'Данс': vals.get('accounts', 0), 'Дебит эргэлт': vals.get('turnover_d', 0), 'Кредит эргэлт': vals.get('turnover_c', 0)} for yr, vals in sorted(tb_stats.items())]
            st.markdown('### 📊 TB summary')
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        if not ml_df.empty:
            st.markdown('### 🤖 TB аномали шинжилгээ')
            show_cols = [c for c in ['year','account_code','account_name','iso_anomaly','zscore_anomaly','turn_anomaly','ensemble_anomaly'] if c in ml_df.columns]
            st.dataframe(ml_df.sort_values(['ensemble_anomaly','year'], ascending=[False, True])[show_cols].head(500), use_container_width=True, hide_index=True)
            _show_dataframe_download(ml_df, 'tb_anomaly_results.csv')
        if not fi.empty:
            st.markdown('### 🔎 TB XAI / Feature importance')
            st.dataframe(fi, use_container_width=True, hide_index=True)
        if not rm_all.empty:
            st.markdown('### 📋 Part1 эрсдэлийн матриц')
            st.dataframe(rm_all.head(300), use_container_width=True, hide_index=True)
        if not mo_all.empty and 'month' in mo_all.columns and 'total_debit_mnt' in mo_all.columns:
            st.markdown('### 📈 Сарын хандлага')
            mo_plot = mo_all.copy()
            mo_plot['total_debit_mnt'] = pd.to_numeric(mo_plot['total_debit_mnt'], errors='coerce').fillna(0)
            fig_mo = px.line(mo_plot, x='month', y='total_debit_mnt', color='year' if 'year' in mo_plot.columns else None, title='Сарын дебит хөдөлгөөн')
            st.plotly_chart(fig_mo, use_container_width=True)
    else:
        st.info('👆 Энд шууд файл оруулах эсвэл 1️⃣ цэсэнд бэлтгэсэн TB/Part1 файлаа ашиглаад шинжилгээг эхлүүлнэ үү.')

elif page.startswith("3"):
    st.header("3️⃣ Ерөнхий журналын шинжилгээ")
    st.markdown("Ledger/ЕЖ файлуудын transaction-level ML + XAI шинжилгээ.")

    ej_files = st.file_uploader("📎 Нэмэлт ЕЖ эсвэл ledger файл оруулах", type=['xlsx','csv','gz'], accept_multiple_files=True, key='journal_page_upload')
    if ej_files:
        st.session_state.journal_upload_cache = uploaded_files_to_cache(ej_files)
    elif st.session_state.get('journal_upload_cache'):
        ej_files = cache_to_file_objects(st.session_state.journal_upload_cache)
        st.info(f"💾 Өмнө оруулсан {len(ej_files)} journal файл хадгалагдсан байна.")

    prepared_ledger_files = _cache_files('prepared_ledger_cache')
    journal_inputs = []
    detect_rows_j = []
    for f in prepared_ledger_files:
        journal_inputs.append(('ledger', f, get_year(f.name)))
    for f in ej_files or []:
        ftype, year = detect_file_type(f)
        f.seek(0)
        detect_rows_j.append({'Файл': f.name, 'Төрөл': FILE_TYPE_LABELS.get(ftype, FILE_TYPE_LABELS['unknown'])[0], 'Он': year})
        if ftype in ('ledger', 'edt'):
            journal_inputs.append((ftype, f, year))
    if detect_rows_j:
        st.dataframe(pd.DataFrame(detect_rows_j), use_container_width=True, hide_index=True)

    cj1, cj2 = st.columns(2)
    with cj1:
        j_cont = st.slider('🎯 Аномалийн хувь', 0.01, 0.20, 0.05, 0.01, key='j_cont_work')
    with cj2:
        j_clusters = st.slider('🧩 KMeans кластерын тоо', 2, 20, 8, 1, key='j_clusters_work')

    with st.expander("🏷️ Эрсдэлийн шинжилгээнээс хасах бүлгүүд", expanded=False):
        excl_settings_j = {}
        for tag, rule in EXCL_RULES.items():
            excl_settings_j[tag] = st.checkbox(rule['label'], value=rule.get('default', False), help=rule.get('help',''), key=f'j_excl_{tag}')

    st.caption(f"Бэлэн journal/ledger: {len(journal_inputs)} файл")
    if st.button('🚀 Ерөнхий журналын шинжилгээ эхлүүлэх', type='primary', use_container_width=True, key='run_journal_analysis_main'):
        try:
            frames = []
            ledger_stats_j = {}
            for typ, f, year in journal_inputs:
                if typ == 'ledger':
                    f.seek(0)
                    led_df = read_ledger(f)
                    if not led_df.empty:
                        frames.append(led_df)
                        ledger_stats_j[Path(f.name).name] = {'rows': len(led_df), 'accounts': led_df.get('account_code', pd.Series(dtype=str)).astype(str).nunique()}
                elif typ == 'edt':
                    f.seek(0)
                    edt_df, cnt = process_edt(f, year)
                    if cnt > 0 and not edt_df.empty:
                        frames.append(edt_df)
                        ledger_stats_j[Path(f.name).name] = {'rows': len(edt_df), 'accounts': edt_df.get('account_code', pd.Series(dtype=str)).astype(str).nunique()}
            if not frames:
                raise ValueError('Journal/ledger өгөгдөл олдсонгүй.')
            ledger_sample_j = pd.concat(frames, ignore_index=True)
            ledger_sample_j = clean_for_risk(ledger_sample_j)
            ml_result_j, ml_feats_j, model_summary_j, xai_importance_j = run_txn_ml_ensemble(ledger_sample_j, contamination=j_cont, n_clusters=j_clusters)
            if ml_result_j.empty:
                raise ValueError('ML шинжилгээ хийхэд хангалттай мөр алга.')
            ml_result_j = classify_exclusions(ml_result_j, level='transaction')
            active_tags_j = [k for k, v in excl_settings_j.items() if v]
            ml_show_j = ml_result_j[~ml_result_j['exclusion_tag'].isin(active_tags_j)].copy() if active_tags_j else ml_result_j.copy()
            st.session_state['journal_ai_done'] = True
            st.session_state['journal_ml_result'] = ml_result_j
            st.session_state['journal_ml_show'] = ml_show_j
            st.session_state['journal_model_summary'] = model_summary_j
            st.session_state['journal_xai'] = xai_importance_j
            st.session_state['journal_ledger_stats'] = ledger_stats_j
            st.session_state['journal_error'] = ''
            st.success('✅ Ерөнхий журналын шинжилгээ дууслаа.')
        except Exception as e:
            st.session_state['journal_error'] = str(e)
            st.exception(e)

    if st.session_state.get('journal_error'):
        st.error(st.session_state['journal_error'])

    if st.session_state.get('journal_ai_done', False):
        ml_show_j = st.session_state.get('journal_ml_show', pd.DataFrame())
        model_summary_j = st.session_state.get('journal_model_summary', pd.DataFrame())
        xai_importance_j = st.session_state.get('journal_xai', pd.DataFrame())
        ledger_stats_j = st.session_state.get('journal_ledger_stats', {})

        if ledger_stats_j:
            st.markdown('### 📊 Journal summary')
            st.dataframe(pd.DataFrame([{'Файл': k, 'Мөр': v.get('rows', 0), 'Данс': v.get('accounts', 0)} for k, v in ledger_stats_j.items()]), use_container_width=True, hide_index=True)
        if not model_summary_j.empty:
            st.markdown('### 🤖 Алгоритмын тойм')
            st.dataframe(model_summary_j, use_container_width=True, hide_index=True)
        if not ml_show_j.empty:
            st.markdown('### 🔍 Өндөр эрсдэлтэй journal гүйлгээ')
            show_cols = [c for c in ['transaction_date','account_code','account_name','counterparty_name','amount','ml_vote_count','ml_risk_level','xai_top_feature','transaction_description'] if c in ml_show_j.columns]
            sort_col = 'ml_vote_count' if 'ml_vote_count' in ml_show_j.columns else show_cols[0]
            st.dataframe(ml_show_j.sort_values(sort_col, ascending=False)[show_cols].head(1000), use_container_width=True, hide_index=True)
            _show_dataframe_download(ml_show_j, 'journal_ml_results.csv')
        if not xai_importance_j.empty:
            st.markdown('### 🧠 XAI / Feature importance')
            st.dataframe(xai_importance_j, use_container_width=True, hide_index=True)
            fig_fi = px.bar(xai_importance_j.head(15), x='importance', y='feature', orientation='h', title='Top XAI features')
            st.plotly_chart(fig_fi, use_container_width=True)
    else:
        st.info('👆 1️⃣ цэсэнд бэлтгэсэн ledger/ЕЖ файлаа ашиглах эсвэл энд шууд оруулж шинжилгээг эхлүүлнэ үү.')

else:
    st.header("4️⃣ Материаллаг байдлын тооцоо")
    st.markdown("TB файл дээр суурилсан материаллаг байдлын хуваарилалт.")

    mat_files = st.file_uploader("📎 Нэмэлт TB файл оруулах", type=['xlsx'], accept_multiple_files=True, key='mat_files_work')
    tb_inputs = _cache_files('prepared_tb_cache')
    for f in mat_files or []:
        ftype, year = detect_file_type(f)
        f.seek(0)
        if ftype == 'tb_std':
            tb_inputs.append(f)
        elif ftype == 'raw_tb':
            buf, tb_sum = process_raw_tb(f)
            if tb_sum is not None and not tb_sum.empty:
                bio = io.BytesIO(buf.getvalue()); bio.name = f'TB_standardized_{year}_{Path(f.name).stem}.xlsx'
                tb_inputs.append(bio)

    total_mat = st.number_input('Нийт материаллаг байдлын дүн', min_value=0.0, value=1000000.0, step=100000.0)
    if st.button('📐 Материаллаг байдлыг тооцоолох', type='primary', use_container_width=True):
        tb_all, _ = load_tb(tb_inputs) if tb_inputs else (pd.DataFrame(), {})
        if tb_all.empty:
            st.warning('TB өгөгдөл олдсонгүй.')
        else:
            d = tb_all.copy()
            base_amt = d['closing_debit'].abs() + d['closing_credit'].abs() if 'closing_debit' in d.columns else d['turnover_debit'].abs() + d['turnover_credit'].abs()
            total_base = base_amt.sum()
            d['materiality_alloc'] = np.where(total_base > 0, total_mat * base_amt / total_base, 0)
            show_cols = [c for c in ['year','account_code','account_name','closing_debit','closing_credit','turnover_debit','turnover_credit','materiality_alloc'] if c in d.columns]
            st.dataframe(d[show_cols].sort_values('materiality_alloc', ascending=False), use_container_width=True, hide_index=True)
            _show_dataframe_download(d[show_cols], 'materiality_by_account.csv')
