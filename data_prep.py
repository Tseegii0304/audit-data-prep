"""
🔍 АУДИТЫН ХОУ — ГҮЙЛГЭЭНИЙ ТҮВШНИЙ ЭРСДЭЛ ИЛРҮҮЛЭГЧ v2.0
ЕДТ файлуудаас гүйлгээ бүрийг шинжилж хэвийн бус гүйлгээг илрүүлнэ.
pip install streamlit pandas numpy scikit-learn plotly openpyxl
streamlit run data_prep.py
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import LabelEncoder, StandardScaler
import warnings, io, re, gzip, math
from datetime import datetime
from collections import Counter
warnings.filterwarnings('ignore')

st.set_page_config(page_title="Гүйлгээний эрсдэл", page_icon="🔍", layout="wide")

# ═══════════════════════════════════════════════════════════
# ТОГТМОЛ & ТУСЛАХ
# ═══════════════════════════════════════════════════════════
EDT_COLUMNS = ['report_year','account_code','account_name','transaction_no','transaction_date',
               'journal_no','document_no','counterparty_name','counterparty_id',
               'transaction_description','debit_mnt','credit_mnt','balance_mnt','month']
ACCT_RE_B = re.compile(r'Данс:\s*\[([^\]]+)\]\s*(.*)')
ACCT_RE_P = re.compile(r'Данс:\s*(\d{3}-\d{2}-\d{2}-\d{3})\s+(.*)')

def safe_float(v):
    if v is None or v == '': return 0.0
    try: return float(v)
    except: return 0.0

def parse_account(text):
    m = ACCT_RE_B.match(text)
    if m: return m.group(1).strip(), m.group(2).strip()
    m = ACCT_RE_P.match(text)
    if m: return m.group(1).strip(), m.group(2).strip()
    return None, None

def get_year(name):
    for y in range(2020, 2030):
        if str(y) in name: return y
    return 2025

# ═══════════════════════════════════════════════════════════
# ФАЙЛ ТАНИХ
# ═══════════════════════════════════════════════════════════
def detect_file_type(f):
    name = f.name.lower()
    year = get_year(f.name)
    if name.endswith('.csv') or name.endswith('.gz'): return 'ledger', year
    if not name.endswith('.xlsx'): return 'unknown', year
    nc = f.name.lower().replace('_',' ').replace('-',' ')
    for kw in ['ерөнхий журнал','ерөнхий дэвтэр','едт','edt','general ledger','general journal','еренхий журнал','journal gc','journal entry','journal entries']:
        if kw in nc: return 'edt', year
    for kw in ['гүйлгээ баланс','гүйлгээ_баланс','гуйлгээ баланс','trial balance','гүйлгэ баланс','journal, tb','journal tb']:
        if kw in nc: return 'raw_tb', year
    if 'tb_standardized' in nc: return 'tb_std', year
    if 'part1' in nc: return 'part1', year
    if 'ledger' in nc: return 'ledger', year
    import openpyxl
    try:
        raw = f.read(); f.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        sheets = wb.sheetnames
        if '02_ACCOUNT_SUMMARY' in sheets:
            has_rm = '04_RISK_MATRIX' in sheets; wb.close()
            return ('part1' if has_rm else 'tb_std'), year
        if '04_RISK_MATRIX' in sheets: wb.close(); return 'part1', year
        ws = wb[sheets[0]]; sample = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            sample.append(row)
            if i >= 300: break
        wb.close()
        for row in sample:
            for cell in row[:6]:
                if cell and str(cell).strip().startswith(('Данс:','Компани:','ЕРӨНХИЙ','Журнал:')): return 'edt', year
        for row in sample:
            if len(row)>=2 and row[1] and re.match(r'\d{3}-\d{2}-\d{2}-\d{3}', str(row[1]).strip()): return 'raw_tb', year
        return 'unknown', year
    except: f.seek(0); return 'unknown', year

FILE_LABELS = {
    'edt': ('📘 ЕДТ / Ерөнхий журнал', 'Гүйлгээний түвшний шинжилгээ хийнэ'),
    'ledger': ('✅ Ledger CSV', 'Гүйлгээний түвшний шинжилгээ хийнэ'),
    'raw_tb': ('📗 ГҮЙЛГЭЭ_БАЛАНС', 'Дансны түвшний мэдээлэл — нэмэлт'),
    'tb_std': ('📊 TB_standardized', 'Нэмэлт мэдээлэл'),
    'part1': ('📈 Part1', 'Нэмэлт мэдээлэл'),
    'unknown': ('❓ Тодорхойгүй', 'Гараар сонгоно уу'),
}

# ═══════════════════════════════════════════════════════════
# ЕДТ УНШИГЧ — ЯМАР Ч ФОРМАТ
# ═══════════════════════════════════════════════════════════
# Баганы нэр таних толь бичиг (монгол + англи)
COL_PATTERNS = {
    'account_code': ['дансны код','дансны дугаар','данс код','account code','account no','account number','acc code','acc no','account','данс №','код'],
    'account_name': ['дансны нэр','данс нэр','account name','acc name','description','нэр'],
    'transaction_date': ['огноо','date','transaction date','тайлбар огноо','txn date','он сар'],
    'debit_mnt': ['дебит','debit','dt','дт','debit amount','дебит дүн'],
    'credit_mnt': ['кредит','credit','ct','кт','credit amount','кредит дүн'],
    'balance_mnt': ['үлдэгдэл','balance','bal','ending balance','эцсийн үлдэгдэл'],
    'counterparty_name': ['харилцагч','counterparty','partner','vendor','customer','нэр','cp name','хар нэр'],
    'transaction_description': ['тайлбар','гүйлгээний утга','утга','description','memo','narration','reference'],
    'journal_no': ['журнал','journal','journal no','jnl'],
    'document_no': ['баримт','document','doc no','баримтын дугаар'],
}

def match_column(header, target_field):
    """Баганы гарчиг нь тухайн талбарт тохирох эсэхийг шалгана."""
    h = str(header).lower().strip()
    for pattern in COL_PATTERNS.get(target_field, []):
        if pattern in h:
            return True
    return False

def auto_map_columns(headers):
    """Баганы гарчигуудаас автоматаар талбар руу нь зурагж (map) өгнө."""
    mapping = {}
    used = set()
    # Эрэмбэлэлт: account_code, debit, credit зэргийг эхэлж олох
    priority = ['account_code','debit_mnt','credit_mnt','transaction_date','account_name',
                'counterparty_name','transaction_description','balance_mnt','journal_no','document_no']
    for field in priority:
        for i, h in enumerate(headers):
            if i in used: continue
            if match_column(h, field):
                mapping[field] = i
                used.add(i)
                break
    return mapping

def process_edt_structured(file_obj, report_year):
    """Стандарт ЕДТ формат: Данс: [xxx-xx-xx-xxx] бүтэцтэй."""
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_out, cur_code, cur_name = [], None, None
    for row in ws.iter_rows(values_only=True):
        c0 = row[0]
        if c0 is None: continue
        s = str(c0).strip()
        if s.startswith('Данс:'):
            code, name = parse_account(s)
            if code: cur_code, cur_name = code, name
            continue
        if any(s.startswith(x) for x in ['Компани:','ЕРӨНХИЙ','Тайлант','Үүсгэсэн','Журнал:','№','Эцсийн','Дт -','Нийт','Эхний','Нээгээд']) or s in ('Валютаар','Төгрөгөөр',''): continue
        try: tx_no = int(float(c0))
        except: continue
        if cur_code is None: continue
        td = row[1] if len(row)>1 else ''
        tx_date = td.strftime('%Y-%m-%d') if isinstance(td, datetime) else (str(td).strip() if td else '')
        rows_out.append({'report_year':str(report_year),'account_code':cur_code,'account_name':cur_name,
            'transaction_no':str(tx_no),'transaction_date':tx_date,
            'journal_no':str(row[5]).strip() if len(row)>5 and row[5] else '',
            'document_no':str(row[6]).strip() if len(row)>6 and row[6] else '',
            'counterparty_name':str(row[3]).strip() if len(row)>3 and row[3] else '',
            'counterparty_id':str(row[4]).strip() if len(row)>4 and row[4] else '',
            'transaction_description':str(row[7]).strip() if len(row)>7 and row[7] else '',
            'debit_mnt':safe_float(row[9]) if len(row)>9 else 0.0,
            'credit_mnt':safe_float(row[11]) if len(row)>11 else 0.0,
            'balance_mnt':safe_float(row[13]) if len(row)>13 else 0.0,
            'month':tx_date[:7] if len(tx_date)>=7 else ''})
    wb.close()
    return rows_out

def process_edt_tabular(file_obj, report_year):
    """Хүснэгт формат: Баганы гарчигтай энгийн Excel."""
    import openpyxl
    raw = file_obj.read(); file_obj.seek(0)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Эхний 20 мөрөөс гарчиг хайх
    all_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        all_rows.append(list(row))
        if i >= 500: break
    wb.close()

    # Гарчиг мөрийг олох (хамгийн олон баганы нэр таарсан мөр)
    best_header_idx = 0
    best_score = 0
    for i, row in enumerate(all_rows[:20]):
        score = 0
        for cell in row:
            if cell is None: continue
            h = str(cell).lower().strip()
            for patterns in COL_PATTERNS.values():
                if any(p in h for p in patterns):
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_header_idx = i

    if best_score < 2:
        return []  # Гарчиг олдсонгүй

    headers = [str(c).strip() if c else f'col_{j}' for j, c in enumerate(all_rows[best_header_idx])]
    col_map = auto_map_columns(headers)

    if 'account_code' not in col_map and 'debit_mnt' not in col_map:
        return []  # Шаардлагатай баганууд олдсонгүй

    rows_out = []
    def _get_val(row, col_map, field, default=''):
        idx = col_map.get(field)
        if idx is None or idx >= len(row): return default
        v = row[idx]
        return str(v).strip() if v is not None else default

    for row in all_rows[best_header_idx+1:]:
        # Хоосон мөр алгасах
        if all(c is None for c in row): continue

        acct_code = _get_val(row, col_map, 'account_code')
        if not acct_code or acct_code in ('None','nan',''): continue

        debit = safe_float(row[col_map['debit_mnt']]) if 'debit_mnt' in col_map and col_map['debit_mnt'] < len(row) else 0.0
        credit = safe_float(row[col_map['credit_mnt']]) if 'credit_mnt' in col_map and col_map['credit_mnt'] < len(row) else 0.0
        if debit == 0 and credit == 0: continue

        td_raw = _get_val(row, col_map, 'transaction_date')
        tx_date = ''
        if td_raw:
            idx = col_map.get('transaction_date')
            if idx is not None and idx < len(row) and isinstance(row[idx], datetime):
                tx_date = row[idx].strftime('%Y-%m-%d')
            else:
                tx_date = td_raw[:10] if len(td_raw) >= 10 else td_raw

        rows_out.append({
            'report_year': str(report_year),
            'account_code': acct_code,
            'account_name': _get_val(row, col_map, 'account_name'),
            'transaction_no': str(len(rows_out)+1),
            'transaction_date': tx_date,
            'journal_no': _get_val(row, col_map, 'journal_no'),
            'document_no': _get_val(row, col_map, 'document_no'),
            'counterparty_name': _get_val(row, col_map, 'counterparty_name'),
            'counterparty_id': '',
            'transaction_description': _get_val(row, col_map, 'transaction_description'),
            'debit_mnt': debit,
            'credit_mnt': credit,
            'balance_mnt': safe_float(row[col_map['balance_mnt']]) if 'balance_mnt' in col_map and col_map['balance_mnt'] < len(row) else 0.0,
            'month': tx_date[:7] if len(tx_date) >= 7 else '',
        })
    return rows_out

def process_edt(file_obj, report_year):
    """Ямар ч форматын ЕДТ/Ерөнхий журнал уншина.
    1) Стандарт ЕДТ формат (Данс: [...]) оролдоно
    2) Амжилтгүй бол хүснэгт форматаар уншина
    """
    file_obj.seek(0)
    # 1-р оролдлого: Стандарт ЕДТ формат
    rows = process_edt_structured(file_obj, report_year)
    if rows:
        return pd.DataFrame(rows), len(rows)

    # 2-р оролдлого: Хүснэгт формат (баганы гарчигтай)
    file_obj.seek(0)
    rows = process_edt_tabular(file_obj, report_year)
    if rows:
        return pd.DataFrame(rows), len(rows)

    # 3-р оролдлого: pandas-аар шууд уншиж баганы нэрээр таних
    file_obj.seek(0)
    try:
        df = pd.read_excel(file_obj)
        col_map = auto_map_columns(df.columns.tolist())
        if 'debit_mnt' in col_map or 'credit_mnt' in col_map:
            rename = {}
            for field, idx in col_map.items():
                rename[df.columns[idx]] = field
            df = df.rename(columns=rename)
            df['report_year'] = str(report_year)
            for c in EDT_COLUMNS:
                if c not in df.columns:
                    df[c] = '' if c in ('account_code','account_name','transaction_description','counterparty_name') else 0
            df['debit_mnt'] = pd.to_numeric(df.get('debit_mnt',0), errors='coerce').fillna(0)
            df['credit_mnt'] = pd.to_numeric(df.get('credit_mnt',0), errors='coerce').fillna(0)
            df = df[(df['debit_mnt']!=0)|(df['credit_mnt']!=0)]
            if 'transaction_date' in df.columns:
                df['month'] = df['transaction_date'].astype(str).str[:7]
            else:
                df['month'] = ''
            return df[EDT_COLUMNS], len(df)
    except: pass

    return pd.DataFrame(columns=EDT_COLUMNS), 0

def read_ledger(f):
    raw = f.read(); f.seek(0)
    if raw[:2]==b'\x1f\x8b': return pd.read_csv(io.StringIO(gzip.decompress(raw).decode('utf-8')), dtype={'account_code':str})
    return pd.read_csv(io.BytesIO(raw), dtype={'account_code':str})

# ═══════════════════════════════════════════════════════════
# 🧠 ГҮЙЛГЭЭНИЙ ТҮВШНИЙ ШИНЖ ЧАНАР ҮҮСГЭХ
# ═══════════════════════════════════════════════════════════
def engineer_transaction_features(df):
    """ЕДТ гүйлгээ бүрээс шинж чанарууд үүсгэнэ."""
    d = df.copy()
    # Дутуу баганууд нэмэх
    for c in ['debit_mnt','credit_mnt','account_code','account_name','counterparty_name','transaction_description','transaction_date']:
        if c not in d.columns:
            d[c] = '' if c in ('account_code','account_name','counterparty_name','transaction_description','transaction_date') else 0
    d['debit_mnt'] = pd.to_numeric(d['debit_mnt'], errors='coerce').fillna(0)
    d['credit_mnt'] = pd.to_numeric(d['credit_mnt'], errors='coerce').fillna(0)
    d['account_code'] = d['account_code'].astype(str).fillna('000')
    d['account_name'] = d['account_name'].astype(str).fillna('')
    d['counterparty_name'] = d['counterparty_name'].astype(str).fillna('')
    d['transaction_description'] = d['transaction_description'].astype(str).fillna('')
    d['transaction_date'] = d['transaction_date'].astype(str).fillna('')

    # 1. Гүйлгээний дүн
    d['amount'] = d['debit_mnt'].abs() + d['credit_mnt'].abs()
    d['log_amount'] = np.log1p(d['amount'])
    d['is_debit'] = (d['debit_mnt'] > 0).astype(int)

    # 2. Дансны ангилал
    d['acct_cat'] = d['account_code'].astype(str).str[:3]
    le = LabelEncoder()
    d['acct_cat_num'] = le.fit_transform(d['acct_cat'].fillna('000'))

    # 3. Бенфордын хуулийн шинж — эхний орон
    def first_digit(x):
        x = abs(x)
        if x < 1: return 0
        return int(str(int(x))[0])
    d['benford_digit'] = d['amount'].apply(first_digit)
    # Хүлээгдэж буй тархалтаас хазайлт
    benford_expected = {1:0.301, 2:0.176, 3:0.125, 4:0.097, 5:0.079, 6:0.067, 7:0.058, 8:0.051, 9:0.046}
    actual_freq = d[d['benford_digit']>0]['benford_digit'].value_counts(normalize=True)
    d['benford_deviation'] = d['benford_digit'].map(lambda x: abs(actual_freq.get(x, 0) - benford_expected.get(x, 0)) if x > 0 else 0)

    # 4. Тоон шинж — бүхэл тоо, тэгстэй тоо
    d['is_round_1000'] = ((d['amount'] >= 1000) & (d['amount'] % 1000 == 0)).astype(int)
    d['is_round_1M'] = ((d['amount'] >= 1_000_000) & (d['amount'] % 1_000_000 == 0)).astype(int)
    d['round_score'] = d['is_round_1000'] + d['is_round_1M']

    # 5. Данс доторх хэвийн бус дүн (z-score)
    acct_stats = d.groupby('account_code')['amount'].agg(['mean','std']).fillna(0)
    acct_stats.columns = ['acct_mean', 'acct_std']
    d = d.merge(acct_stats, on='account_code', how='left')
    d['amount_zscore'] = np.where(d['acct_std'] > 0, (d['amount'] - d['acct_mean']) / d['acct_std'], 0)
    d['amount_zscore'] = d['amount_zscore'].clip(-10, 10)

    # 6. Харилцагчийн давтамж (ховор харилцагч = эрсдэлтэй)
    cp_freq = d['counterparty_name'].value_counts()
    d['cp_frequency'] = d['counterparty_name'].map(cp_freq).fillna(0)
    d['cp_is_rare'] = (d['cp_frequency'] <= 3).astype(int)

    # 7. Данс-харилцагч хосын ховор байдал
    d['acct_cp_pair'] = d['account_code'] + '|' + d['counterparty_name'].fillna('')
    pair_freq = d['acct_cp_pair'].value_counts()
    d['pair_frequency'] = d['acct_cp_pair'].map(pair_freq).fillna(0)
    d['pair_is_rare'] = (d['pair_frequency'] <= 2).astype(int)

    # 8. Гүйлгээний тайлбарын урт ба ховор байдал
    d['desc_len'] = d['transaction_description'].fillna('').str.len()
    d['desc_empty'] = (d['desc_len'] == 0).astype(int)

    # ═══════════════════════════════════════════════════════
    # 11. ГҮЙЛГЭЭНИЙ ТАЙЛБАР ↔ ДАНСНЫ НЭР ТУЛГАЛТ
    # ═══════════════════════════════════════════════════════
    d['desc_acct_mismatch'] = 0
    d['name_desc_overlap'] = 0.0
    d['name_desc_no_overlap'] = 0
    d['desc_unusual_ratio'] = 0.0
    d['acct_direction_mismatch'] = 0
    try:
        desc = d['transaction_description'].str.lower().str.strip()
        acct_name = d['account_name'].str.lower().str.strip()

        def extract_keywords(text):
            if not text: return set()
            words = re.findall(r'[а-яөүё\w]{3,}', text)
            stop = {'дансны','данс','нийт','бусад','зардал','орлого','төлбөр','хөрөнгө',
                    'тооцоо','тооцооны','бүртгэл','дүн','төгрөг','сая','мянга','нэг',
                    'хоёр','гурав','дөрөв','тав','зургаа','долоо','найм','ес','арав',
                    'оны','онд','сарын','сард','өдрийн','журнал','гүйлгээ','баримт'}
            return set(w for w in words if w not in stop and len(w) >= 3)

        acct_typical_words = {}
        for code in d['account_code'].unique():
            all_descs = ' '.join(d.loc[d['account_code']==code, 'transaction_description'].str.lower())
            wc = Counter(re.findall(r'[а-яөүё\w]{3,}', all_descs))
            acct_typical_words[code] = set(w for w, c in wc.items() if c >= 3 and len(w) >= 3)

        def _desc_mismatch(code, tx_desc):
            tx = str(tx_desc).lower() if tx_desc else ''
            if not tx or code not in acct_typical_words or not acct_typical_words[code]: return 0
            return 0 if len(set(re.findall(r'[а-яөүё\w]{3,}', tx)) & acct_typical_words[code]) > 0 else 1
        d['desc_acct_mismatch'] = [_desc_mismatch(c, t) for c, t in zip(d['account_code'], d['transaction_description'])]

        def _name_overlap(aname, tdesc):
            a, t = str(aname).lower() if aname else '', str(tdesc).lower() if tdesc else ''
            if not a or not t: return 0
            nk, dk = extract_keywords(a), extract_keywords(t)
            if not nk or not dk: return 0
            return len(nk & dk) / max(len(nk), 1)
        d['name_desc_overlap'] = [_name_overlap(a, t) for a, t in zip(d['account_name'], d['transaction_description'])]
        d['name_desc_no_overlap'] = (d['name_desc_overlap'] == 0).astype(int)

        def _desc_rarity(code, tdesc):
            tx = str(tdesc).lower() if tdesc else ''
            if not tx or code not in acct_typical_words or not acct_typical_words[code]: return 0
            tx_words = set(re.findall(r'[а-яөүё\w]{3,}', tx))
            if not tx_words: return 0
            return len(tx_words - acct_typical_words[code]) / max(len(tx_words), 1)
        d['desc_unusual_ratio'] = [_desc_rarity(c, t) for c, t in zip(d['account_code'], d['transaction_description'])]

        acct_first = d['account_code'].str[0]
        d.loc[(acct_first=='1') & (d['credit_mnt']>0) & (d['debit_mnt']==0), 'acct_direction_mismatch'] = 1
        d.loc[(acct_first=='2') & (d['debit_mnt']>0) & (d['credit_mnt']==0), 'acct_direction_mismatch'] = 1
        d.loc[(acct_first=='5') & (d['debit_mnt']>0) & (d['credit_mnt']==0), 'acct_direction_mismatch'] = 1
        d.loc[(acct_first.isin(['6','7','8'])) & (d['credit_mnt']>0) & (d['debit_mnt']==0), 'acct_direction_mismatch'] = 1
    except Exception:
        pass

    # 9. Цаг хугацааны шинж
    d['day'] = pd.to_numeric(d['transaction_date'].str[8:10], errors='coerce').fillna(15)
    d['month_num'] = pd.to_numeric(d['transaction_date'].str[5:7], errors='coerce').fillna(6)
    d['is_month_end'] = (d['day'] >= 28).astype(int)
    d['is_year_end'] = (d['month_num'] == 12).astype(int)

    # 10. Давхардал (ижил данс + ижил дүн + ижил өдөр)
    d['dup_key'] = d['account_code'] + '|' + d['amount'].astype(str) + '|' + d['transaction_date']
    dup_counts = d['dup_key'].value_counts()
    d['dup_count'] = d['dup_key'].map(dup_counts).fillna(1)
    d['is_duplicate'] = (d['dup_count'] > 1).astype(int)

    return d

# ═══════════════════════════════════════════════════════════
# 🔍 ГҮЙЛГЭЭНИЙ АНОМАЛИ ИЛРҮҮЛЭЛТ
# ═══════════════════════════════════════════════════════════
def run_transaction_anomaly(df, contamination=0.05):
    """Гүйлгээний түвшинд хэвийн бус байдлыг илрүүлнэ."""
    features = ['log_amount','acct_cat_num','benford_deviation','round_score',
                'amount_zscore','cp_is_rare','pair_is_rare','desc_empty',
                'is_month_end','is_year_end','is_duplicate','is_debit',
                'desc_acct_mismatch','name_desc_no_overlap','desc_unusual_ratio','acct_direction_mismatch']
    # Дутуу feature нэмэх
    for f in features:
        if f not in df.columns:
            df[f] = 0

    X = df[features].fillna(0).replace([np.inf, -np.inf], 0).astype(float)

    # Isolation Forest
    iso = IsolationForest(contamination=contamination, random_state=42, n_estimators=200, n_jobs=1)
    df['anomaly_score'] = -iso.fit(X).score_samples(X)
    df['is_anomaly'] = (iso.predict(X) == -1).astype(int)

    # Z-score нэмэлт шалгалт
    scaler = StandardScaler()
    z = np.abs(scaler.fit_transform(X))
    df['max_zscore'] = z.max(axis=1)
    df['zscore_flag'] = (df['max_zscore'] > 2.5).astype(int)

    # Нэгдсэн эрсдэлийн оноо
    df['risk_score'] = (
        df['is_anomaly'] * 3 +          # IF аномали
        df['zscore_flag'] * 2 +          # Z-score хазайлт
        df['is_duplicate'] * 2 +         # Давхардал
        df['cp_is_rare'] * 1 +           # Ховор харилцагч
        df['pair_is_rare'] * 1 +         # Ховор данс-харилцагч хос
        df['is_round_1M'] * 1 +          # Тэгс тоо (сая)
        (df['amount_zscore'].abs() > 3).astype(int) * 2 +  # Данс доторх хэт их дүн
        df['desc_empty'] * 1 +           # Тайлбаргүй
        df['desc_acct_mismatch'] * 2 +   # Тайлбар дансны ердийн хэв маягаас зөрсөн
        df['name_desc_no_overlap'] * 1 + # Дансны нэр ↔ тайлбар огт давхцахгүй
        (df['desc_unusual_ratio'] > 0.7).astype(int) * 2 + # Тайлбарын 70%+ нь тухайн дансанд ер бусын
        df['acct_direction_mismatch'] * 2 # Дансны төрөл ↔ дебит/кредит чиглэл зөрсөн
    )

    # Эрсдэлийн түвшин
    df['risk_level'] = pd.cut(df['risk_score'],
        bins=[-1, 3, 7, 12, 100],
        labels=['🟢 Бага', '🟡 Дунд', '🟠 Өндөр', '🔴 Маш өндөр'])

    return df, features

# ═══════════════════════════════════════════════════════════
# ИНТЕРФЕЙС
# ═══════════════════════════════════════════════════════════
st.markdown('<h1 style="text-align:center;color:#1565c0">🔍 Гүйлгээний түвшний эрсдэл илрүүлэгч</h1>', unsafe_allow_html=True)
st.markdown("""
<div style="background:#E3F2FD; padding:15px; border-radius:10px; border-left:5px solid #1565C0; margin-bottom:20px;">
    <b>📂 ЕДТ / Ерөнхий журнал файлуудаа оруулаарай</b><br>
    <span style="color:#555; font-size:13px;">
    Гүйлгээ бүрийг дансны код, нэр, харилцагч, дүн, тайлбараар нь шинжилж хэвийн бус гүйлгээг илрүүлнэ.
    </span>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader("📎 ЕДТ / Ledger файлуудаа оруулна уу", type=['xlsx','csv','gz'], accept_multiple_files=True, key='txn_files')

if uploaded:
    # ── Файл таних & уншиж нэгтгэх ──
    all_txn = []
    for f in uploaded:
        ftype, year = detect_file_type(f); f.seek(0)
        lb, desc = FILE_LABELS.get(ftype, FILE_LABELS['unknown'])
        st.write(f"**{f.name}** → {lb}")

        if ftype == 'edt':
            with st.spinner(f"📘 {f.name} уншиж байна..."):
                df_e, cnt = process_edt(f, year)
            if cnt > 0:
                st.success(f"✅ {cnt:,} гүйлгээ уншлаа")
                all_txn.append(df_e)
            else:
                st.warning(f"⚠️ Гүйлгээ уншигдсангүй")
        elif ftype == 'ledger':
            with st.spinner(f"📄 {f.name} уншиж байна..."):
                df_l = read_ledger(f)
                df_l['report_year'] = str(year)
            st.success(f"✅ {len(df_l):,} гүйлгээ уншлаа")
            all_txn.append(df_l)
        elif ftype == 'unknown':
            choice = st.selectbox(f"**{f.name}** — Төрөл сонгох:", ['skip','edt'],
                format_func=lambda x: {'skip':'❌ Алгасах','edt':'📘 ЕДТ гэж уншаарай'}[x], key=f"m_{f.name}")
            if choice == 'edt':
                f.seek(0)
                df_e, cnt = process_edt(f, year)
                if cnt > 0: all_txn.append(df_e); st.success(f"✅ {cnt:,} гүйлгээ")
                else: st.warning("⚠️ Гүйлгээ уншигдсангүй")
        else:
            st.info(f"ℹ️ {ftype} файл — гүйлгээний шинжилгээнд ашиглагдахгүй")

    if all_txn:
        txn = pd.concat(all_txn, ignore_index=True)
        st.markdown("---")
        st.markdown(f"### 📊 Нийт: **{len(txn):,}** гүйлгээ, **{txn['account_code'].nunique():,}** данс")

        # ── Тохиргоо ──
        st.markdown("""
        <div style="background:#F5F5F5; padding:12px; border-radius:8px; margin-bottom:10px;">
        <b>⚙️ Шинжилгээний тохиргоо</b> — параметрүүдийг өөрчилж илрүүлэлтийн мэдрэмжийг тохируулна
        </div>
        """, unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            cont = st.slider("🎯 IF contamination — Хэвийн бус гүйлгээний хувь (%)", 1, 20, 5, 1,
                help="Isolation Forest (Тусгаарлалтын ой) алгоритм нийт гүйлгээний хэдэн хувийг хэвийн бус гэж үзэх. "
                     "5% = зөвхөн хамгийн сэжигтэй 5%-ийг илрүүлнэ. "
                     "20% = илүү олон гүйлгээг шалгана (илүү өргөн хүрээ).") / 100
        with c2:
            max_rows = st.selectbox("📊 Шинжлэх гүйлгээний хязгаар", [50000, 100000, 500000, 1000000, len(txn)],
                format_func=lambda x: f"{x:,}" if x < len(txn) else f"Бүгд ({len(txn):,})",
                help="Том файлын хувьд бүх гүйлгээг шинжлэхэд удаан байж болно. "
                     "Түүвэрлэлт хийвэл хурдан шинжлэнэ, гэхдээ бүх гүйлгээг хамрахгүй.")

        if st.button("🚀 Гүйлгээний эрсдэл илрүүлэх", type="primary", use_container_width=True):
            # Хэрэв хэт олон мөр бол түүвэрлэнэ
            if len(txn) > max_rows:
                st.info(f"🔄 {len(txn):,} → {max_rows:,} гүйлгээ түүвэрлэж байна...")
                txn_sample = txn.sample(n=max_rows, random_state=42)
            else:
                txn_sample = txn

            with st.spinner("🔧 Шинж чанар үүсгэж байна..."):
                txn_feat = engineer_transaction_features(txn_sample)

            with st.spinner("🤖 Хэвийн бус гүйлгээ илрүүлж байна..."):
                result, feat_names = run_transaction_anomaly(txn_feat, cont)

            st.session_state['txn_result'] = result
            st.session_state['feat_names'] = feat_names
            st.session_state['txn_done'] = True

    # ── ҮР ДҮН ──
    if st.session_state.get('txn_done'):
        result = st.session_state['txn_result']
        feat_names = st.session_state['feat_names']

        n_anomaly = result['is_anomaly'].sum()
        n_total = len(result)

        st.markdown("---")
        st.success(f"✅ Шинжилгээ дууслаа: {n_total:,} гүйлгээ, {n_anomaly:,} хэвийн бус ({n_anomaly/n_total*100:.1f}%)")

        tabs = st.tabs(["📊 Тойм", "🔴 Хэвийн бус гүйлгээ", "🏷️ Дансаар", "👤 Харилцагчаар", "📈 Шинж чанар", "📥 Татах"])

        # ── TAB 1: Тойм ──
        with tabs[0]:
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Нийт гүйлгээ", f"{n_total:,}")
            c2.metric("Хэвийн бус", f"{n_anomaly:,}", delta=f"{n_anomaly/n_total*100:.1f}%", delta_color="inverse")
            c3.metric("Давхардсан", f"{result['is_duplicate'].sum():,}")
            c4.metric("Ховор харилцагч", f"{result['cp_is_rare'].sum():,}")

            # Эрсдэлийн түвшний тархалт
            risk_dist = result['risk_level'].value_counts().reindex(['🟢 Бага','🟡 Дунд','🟠 Өндөр','🔴 Маш өндөр']).fillna(0)
            fig = px.bar(x=risk_dist.index, y=risk_dist.values, color=risk_dist.index,
                color_discrete_map={'🟢 Бага':'#4CAF50','🟡 Дунд':'#FFC107','🟠 Өндөр':'#FF9800','🔴 Маш өндөр':'#F44336'},
                labels={'x':'Эрсдэлийн түвшин','y':'Гүйлгээний тоо'}, title="Эрсдэлийн түвшний тархалт")
            fig.update_layout(height=350, showlegend=False); st.plotly_chart(fig, use_container_width=True)

            # Бенфордын хууль
            st.subheader("Бенфордын хуулийн шалгалт")
            benford_exp = {1:30.1,2:17.6,3:12.5,4:9.7,5:7.9,6:6.7,7:5.8,8:5.1,9:4.6}
            actual = result[result['benford_digit']>0]['benford_digit'].value_counts(normalize=True).sort_index()*100
            ben_df = pd.DataFrame({'Орон':range(1,10), 'Хүлээгдэж буй (%)': [benford_exp[i] for i in range(1,10)],
                'Бодит (%)': [actual.get(i,0) for i in range(1,10)]})
            ben_df['Зөрүү'] = (ben_df['Бодит (%)'] - ben_df['Хүлээгдэж буй (%)']).round(2)
            fig_b = go.Figure()
            fig_b.add_trace(go.Bar(x=ben_df['Орон'], y=ben_df['Хүлээгдэж буй (%)'], name='Бенфорд (хүлээгдэж буй)', marker_color='#90CAF9'))
            fig_b.add_trace(go.Bar(x=ben_df['Орон'], y=ben_df['Бодит (%)'], name='Бодит тархалт', marker_color='#1565C0'))
            fig_b.update_layout(barmode='group', height=300, title="Эхний оронгийн тархалт"); st.plotly_chart(fig_b, use_container_width=True)

        # ── TAB 2: Хэвийн бус гүйлгээний жагсаалт ──
        with tabs[1]:
            st.subheader("🔴 Хэвийн бус гүйлгээний жагсаалт")
            dp_years = sorted(result['report_year'].dropna().unique().tolist()) if 'report_year' in result.columns else []
            fc1, fc2 = st.columns(2)
            with fc1:
                risk_filter = st.selectbox("Эрсдэлийн түвшин:", ['Бүгд','🔴 Маш өндөр','🟠 Өндөр','🟡 Дунд'], key='dp_risk_f')
            with fc2:
                year_filter = st.selectbox("Он:", ['Бүгд'] + [str(y) for y in dp_years], key='dp_year_f')
            anom = result[result['is_anomaly']==1].copy()
            if risk_filter != 'Бүгд':
                anom = anom[anom['risk_level']==risk_filter]
            if year_filter != 'Бүгд' and 'report_year' in anom.columns:
                anom = anom[anom['report_year'].astype(str)==year_filter]
            show_cols = ['risk_level','risk_score','account_code','account_name','counterparty_name',
                         'transaction_date','debit_mnt','credit_mnt','transaction_description',
                         'amount_zscore','is_duplicate','cp_is_rare',
                         'desc_acct_mismatch','name_desc_no_overlap','acct_direction_mismatch']
            anom_show = anom[[c for c in show_cols if c in anom.columns]].sort_values('risk_score', ascending=False)
            st.write(f"Нийт: **{len(anom_show):,}** гүйлгээ")
            st.dataframe(anom_show, use_container_width=True, hide_index=True, height=500)
            st.download_button("📥 CSV татах", anom_show.to_csv(index=False).encode('utf-8-sig'), "anomaly_transactions.csv")

        # ── TAB 3: Дансаар нэгтгэл ──
        with tabs[2]:
            st.subheader("🏷️ Дансаар нэгтгэсэн эрсдэлийн тойм")
            dp_years3 = sorted(result['report_year'].dropna().unique().tolist()) if 'report_year' in result.columns else []
            year_f3 = st.selectbox("Он:", ['Бүгд'] + [str(y) for y in dp_years3], key='dp_acct_year')
            res_f3 = result.copy()
            if year_f3 != 'Бүгд' and 'report_year' in res_f3.columns:
                res_f3 = res_f3[res_f3['report_year'].astype(str)==year_f3]
            acct_risk = res_f3.groupby(['account_code','account_name']).agg(
                total_txn=('amount','count'), anomaly_txn=('is_anomaly','sum'),
                total_amount=('amount','sum'), avg_risk=('risk_score','mean'),
                max_risk=('risk_score','max')
            ).reset_index()
            acct_risk['anomaly_pct'] = (acct_risk['anomaly_txn']/acct_risk['total_txn']*100).round(1)
            acct_risk = acct_risk.sort_values('anomaly_txn', ascending=False)
            st.write(f"Нийт: **{len(acct_risk):,}** данс")
            st.dataframe(acct_risk.head(50), use_container_width=True, hide_index=True)
            st.download_button("📥 Дансны жагсаалт CSV", acct_risk.to_csv(index=False).encode('utf-8-sig'), "account_risk.csv", key='dp_dl_acct')
            top20 = acct_risk.head(20)
            if len(top20) > 0:
                fig_a = px.bar(top20, x='anomaly_txn', y='account_name', orientation='h',
                    color='avg_risk', color_continuous_scale='Reds',
                    title='Топ 20 — хэвийн бус гүйлгээ хамгийн олонтой данс',
                    labels={'anomaly_txn':'Хэвийн бус гүйлгээний тоо','account_name':'Данс','avg_risk':'Дундаж эрсдэл'})
                fig_a.update_layout(height=500, yaxis={'categoryorder':'total ascending'}); st.plotly_chart(fig_a, use_container_width=True)

        # ── TAB 4: Харилцагчаар ──
        with tabs[3]:
            st.subheader("👤 Харилцагчаар нэгтгэсэн эрсдэлийн тойм")
            dp_years4 = sorted(result['report_year'].dropna().unique().tolist()) if 'report_year' in result.columns else []
            year_f4 = st.selectbox("Он:", ['Бүгд'] + [str(y) for y in dp_years4], key='dp_cp_year')
            res_f4 = result.copy()
            if year_f4 != 'Бүгд' and 'report_year' in res_f4.columns:
                res_f4 = res_f4[res_f4['report_year'].astype(str)==year_f4]
            cp_risk = res_f4[res_f4['counterparty_name']!=''].groupby('counterparty_name').agg(
                total_txn=('amount','count'), anomaly_txn=('is_anomaly','sum'),
                total_amount=('amount','sum'), accounts_used=('account_code','nunique'),
                avg_risk=('risk_score','mean')
            ).reset_index()
            cp_risk['anomaly_pct'] = (cp_risk['anomaly_txn']/cp_risk['total_txn']*100).round(1)
            cp_risk = cp_risk.sort_values('anomaly_txn', ascending=False)
            st.write(f"Нийт: **{len(cp_risk):,}** харилцагч")
            st.dataframe(cp_risk.head(50), use_container_width=True, hide_index=True)
            st.download_button("📥 Харилцагчийн жагсаалт CSV", cp_risk.to_csv(index=False).encode('utf-8-sig'), "counterparty_risk.csv", key='dp_dl_cp')
            top20_cp = cp_risk.head(20)
            if len(top20_cp) > 0:
                fig_cp = px.bar(top20_cp, x='anomaly_txn', y='counterparty_name', orientation='h',
                    color='accounts_used', color_continuous_scale='Blues',
                    title='Топ 20 — хэвийн бус гүйлгээтэй харилцагч',
                    labels={'anomaly_txn':'Хэвийн бус','counterparty_name':'Харилцагч','accounts_used':'Ашигласан дансны тоо'})
                fig_cp.update_layout(height=500, yaxis={'categoryorder':'total ascending'}); st.plotly_chart(fig_cp, use_container_width=True)

        # ── TAB 5: Шинж чанарын шинжилгээ ──
        with tabs[4]:
            st.subheader("📈 Шинж чанарын ач холбогдол")
            # Anomaly vs Normal comparison
            feat_desc = {'log_amount':'Гүйлгээний дүн (логарифм)','acct_cat_num':'Дансны ангилал',
                'benford_deviation':'Бенфордын хуулиас хазайлт','round_score':'Тэгс тоон оноо',
                'amount_zscore':'Данс доторх хэвийн бус дүн','cp_is_rare':'Ховор харилцагч',
                'pair_is_rare':'Ховор данс-харилцагч хос','desc_empty':'Тайлбаргүй гүйлгээ',
                'is_month_end':'Сарын эцэс','is_year_end':'Жилийн эцэс',
                'is_duplicate':'Давхардсан гүйлгээ','is_debit':'Дебит гүйлгээ',
                'desc_acct_mismatch':'⚠️ Тайлбар ↔ дансны ердийн хэв маяг зөрсөн',
                'name_desc_no_overlap':'⚠️ Дансны нэр ↔ тайлбар огт давхцахгүй',
                'desc_unusual_ratio':'⚠️ Тухайн дансанд ер бусын тайлбар',
                'acct_direction_mismatch':'⚠️ Дансны төрөл ↔ дебит/кредит чиглэл зөрсөн'}
            comp_rows = []
            for f in feat_names:
                norm_mean = result[result['is_anomaly']==0][f].mean()
                anom_mean = result[result['is_anomaly']==1][f].mean()
                diff = anom_mean - norm_mean
                comp_rows.append({'Шинж чанар':feat_desc.get(f,f), 'Код':f,
                    'Хэвийн (дундаж)':f"{norm_mean:.3f}", 'Хэвийн бус (дундаж)':f"{anom_mean:.3f}",
                    'Зөрүү':f"{diff:+.3f}"})
            st.dataframe(pd.DataFrame(comp_rows), use_container_width=True, hide_index=True)

            # Scatter
            st.plotly_chart(px.scatter(result.sample(min(5000, len(result))), x='log_amount', y='amount_zscore',
                color=result.sample(min(5000, len(result)))['is_anomaly'].map({0:'Хэвийн',1:'Хэвийн бус'}),
                color_discrete_map={'Хэвийн':'#90caf9','Хэвийн бус':'#c62828'}, opacity=0.4,
                title='Гүйлгээний дүн ↔ Данс доторх хазайлт', height=400,
                labels={'log_amount':'Гүйлгээний дүн (log)','amount_zscore':'Данс доторх z-score'}
            ), use_container_width=True)

        # ── TAB 6: Татах ──
        with tabs[5]:
            st.subheader("📥 Үр дүн татах")
            # Бүх гүйлгээ эрсдэлийн оноотой
            dl_cols = ['report_year','account_code','account_name','counterparty_name','transaction_date',
                       'debit_mnt','credit_mnt','transaction_description','risk_score','risk_level',
                       'is_anomaly','is_duplicate','amount_zscore','cp_is_rare',
                       'desc_acct_mismatch','name_desc_no_overlap','desc_unusual_ratio','acct_direction_mismatch']
            dl_df = result[[c for c in dl_cols if c in result.columns]].copy()
            c1, c2 = st.columns(2)
            c1.download_button("📥 Бүх гүйлгээ (эрсдэлийн оноотой)", dl_df.to_csv(index=False).encode('utf-8-sig'),
                "all_transactions_scored.csv", "text/csv")
            anom_only = dl_df[dl_df['is_anomaly']==1].sort_values('risk_score', ascending=False)
            c2.download_button(f"📥 Зөвхөн хэвийн бус ({len(anom_only):,})", anom_only.to_csv(index=False).encode('utf-8-sig'),
                "anomaly_transactions.csv", "text/csv")

            st.markdown("""
            ---
            **Шинж чанарын тайлбар:**
            | Шинж | Тайлбар |
            |------|---------|
            | `risk_score` | Нэгдсэн эрсдэлийн оноо (0-20+) |
            | `risk_level` | 🟢 Бага / 🟡 Дунд / 🟠 Өндөр / 🔴 Маш өндөр |
            | `is_anomaly` | Тусгаарлалтын ойн дүн (1=хэвийн бус) |
            | `amount_zscore` | Тухайн дансны дундажаас хэр их зөрсөн |
            | `is_duplicate` | Давхардсан гүйлгээ (ижил данс+дүн+огноо) |
            | `cp_is_rare` | Ховор харилцагч (≤3 удаа гарсан) |
            | `desc_acct_mismatch` | ⚠️ Тайлбар нь тухайн дансны ердийн хэв маягаас зөрсөн |
            | `name_desc_no_overlap` | ⚠️ Дансны нэр ↔ гүйлгээний тайлбар огт давхцахгүй |
            | `desc_unusual_ratio` | Тайлбарын хэдэн хувь нь тухайн дансанд ер бусын (0-1) |
            | `acct_direction_mismatch` | ⚠️ Дансны төрөл ↔ дебит/кредит чиглэл зөрсөн |
            """)

elif not uploaded:
    st.markdown("""
    ---
    ### 📋 Илрүүлэх шинж чанарууд:

    | # | Шинж | Тайлбар | ISA холбоос |
    |---|------|---------|------------|
    | 1 | **Гүйлгээний дүн** | Данс доторх хэвийн бус дүн (z-score) | ISA 520 |
    | 2 | **Бенфордын хууль** | Эхний оронгийн тархалтын хазайлт | ISA 240 |
    | 3 | **Тэгс тоо** | Бүхэл/тэгс дүнтэй гүйлгээ | ISA 240 |
    | 4 | **Ховор харилцагч** | ≤3 удаа гарсан харилцагч | ISA 550 |
    | 5 | **Ховор данс-харилцагч хос** | Ер бусын хослол | ISA 550 |
    | 6 | **Давхардал** | Ижил данс+дүн+огноо | ISA 240 |
    | 7 | **Тайлбаргүй гүйлгээ** | Хоосон тайлбартай | ISA 500 |
    | 8 | **Сарын/жилийн эцэс** | Тайлант хугацааны эцэст хийсэн | ISA 240 |
    | 9 | **Дансны ангилал** | Тодорхой ангиллын эрсдэл | ISA 315 |
    | 10 | **Тайлбар ↔ дансны хэв маяг** | ⚠️ Гүйлгээний тайлбар тухайн дансны ердийн тайлбараас зөрсөн | ISA 500 |
    | 11 | **Дансны нэр ↔ тайлбар** | ⚠️ Дансны нэрийн түлхүүр үг тайлбарт огт байхгүй | ISA 500 |
    | 12 | **Тухайн дансанд ер бусын тайлбар** | ⚠️ Тайлбарын ихэнх үг тухайн дансанд хэзээ ч гараагүй | ISA 315 |
    | 13 | **Дансны төрөл ↔ чиглэл** | ⚠️ Хөрөнгийн дансанд кредит, орлогын дансанд дебит гэх мэт | ISA 240 |
    """)
